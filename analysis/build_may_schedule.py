"""
Генератор Excel с планом мая 2026 (день за днём).
Запуск: python -m analysis.build_may_schedule
Сохраняет: plans/2026_05_may_schedule.xlsx

v2: разделил утро/вечер по строкам, добавил детальный лист «Силовая»,
явно указано «в подъём/на равнине» для ускорений, скорректированы часы.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path("plans/2026_05_may_schedule.xlsx")
OUT.parent.mkdir(parents=True, exist_ok=True)


def save_with_fallback(wb, primary):
    """Если основной файл занят (открыт в Excel) — сохраняем как _v2."""
    try:
        wb.save(primary)
        return primary
    except PermissionError:
        alt = primary.with_name(primary.stem + "_v2" + primary.suffix)
        wb.save(alt)
        return alt

# Тип тренировки → цвет
TYPE_COLORS = {
    "вело":      "BBDEFB",
    "вело-длит": "1976D2",
    "бег":       "C8E6C9",
    "бег-длит":  "388E3C",
    "ст-омв":    "FFCC80",
    "ст-омв-пик":"E65100",
    "ст-тон":    "FFE0B2",
    "контроль":  "FFEB3B",
    "отдых":     "ECEFF1",
    "":          "FFFFFF",
}

WHITE_ON = {"вело-длит", "бег-длит", "ст-омв-пик"}  # тёмные — белый шрифт

# (дата, неделя, день, утро_текст, утро_тип, утро_часы, вечер_текст, вечер_тип, вечер_часы)
# ⚠ Все длительные субботы в мае ограничены 3-3:20 (вело-фокус, без перегрузки)
# Освободившиеся часы размазаны по будням (среда/пятница) и воскресеньям.
PLAN = [
    # W1 хвост (~5.6ч)
    ("2026-05-01", 1, "Пт", "вело 1ч Z2 + 4×30 сек разгон НА РАВНИНЕ",                 "вело",    1.0,  "",                                              "",       0),
    ("2026-05-02", 1, "Сб", "длит. вело 3ч 20м Z1-Z2",                                 "вело-длит",3.33, "",                                              "",       0),
    ("2026-05-03", 1, "Вс", "бег 1ч 15м Z2 (рельеф)",                                  "бег",     1.25, "",                                              "",       0),

    # W2 (~14.1ч): Вело база 1 + СТАРТ блока СТ ОМВ
    ("2026-05-04", 2, "Пн", "бег 1ч Z1 (восст активн.)",                               "бег",     1.0,  "",                                              "",       0),
    ("2026-05-05", 2, "Вт", "вело 1ч 15м Z2",                                          "вело",    1.25, "🏋 СТ ОМВ-НИЗ ЗАЛ 1ч (10м разм. вело + 3 круга + 5м зам)", "ст-омв", 1.0),
    ("2026-05-06", 2, "Ср", "вело 1ч 45м Z2",                                          "вело",    1.75, "",                                              "",       0),
    ("2026-05-07", 2, "Чт", "бег 1ч Z1 + растяжка 15м",                                "бег",     1.0,  "",                                              "",       0),
    ("2026-05-08", 2, "Пт", "вело 1ч 30м Z2 + 4×30 сек В ПОДЪЁМ",                       "вело",    1.5,  "🤸 СТ ОМВ-ВЕРХ УЛИЦА 1ч (10м разм. бег + 3 круга + 5м зам)", "ст-омв", 1.0),
    ("2026-05-09", 2, "Сб", "длит. вело 3ч 20м Z1-Z2",                                 "вело-длит",3.33, "",                                              "",       0),
    ("2026-05-10", 2, "Вс", "бег 2ч 15м Z2 (рельеф) — длинная",                        "бег-длит",2.25, "",                                              "",       0),

    # W3 (~14.9ч): Вело база 2 + СТ ОМВ развивающий
    ("2026-05-11", 3, "Пн", "бег 1ч 15м Z1",                                           "бег",     1.25, "",                                              "",       0),
    ("2026-05-12", 3, "Вт", "вело 1ч 15м Z2",                                          "вело",    1.25, "🏋 СТ ОМВ-НИЗ ЗАЛ 1ч 15м (10м разм. вело + 4 круга + 10м зам)", "ст-омв", 1.25),
    ("2026-05-13", 3, "Ср", "вело 1ч 45м Z2 + 4×20 сек В КОРОТКИЙ ПОДЪЁМ",              "вело",    1.75, "",                                              "",       0),
    ("2026-05-14", 3, "Чт", "бег 1ч Z1 + растяжка",                                    "бег",     1.0,  "",                                              "",       0),
    ("2026-05-15", 3, "Пт", "вело 1ч 30м Z2",                                          "вело",    1.5,  "🤸 СТ ОМВ-ВЕРХ УЛИЦА 1ч 15м (10м разм. бег + 4 круга + 10м зам)", "ст-омв", 1.25),
    ("2026-05-16", 3, "Сб", "длит. вело 3ч 20м Z1-Z2 (рельеф)",                        "вело-длит",3.33, "",                                              "",       0),
    ("2026-05-17", 3, "Вс", "бег 2ч 20м Z2 (рельеф) — длинная",                        "бег-длит",2.33, "",                                              "",       0),

    # W4 (~15.6ч): Нагрузочный ПИК
    ("2026-05-18", 4, "Пн", "бег 1ч 20м Z1 (восст)",                                   "бег",     1.33, "",                                              "",       0),
    ("2026-05-19", 4, "Вт", "вело 1ч 15м Z2",                                          "вело",    1.25, "🏋 СТ ОМВ-НИЗ ЗАЛ 1ч 15м (10м разм. вело + 5 кругов ПИК + 10м зам)", "ст-омв-пик",1.25),
    ("2026-05-20", 4, "Ср", "вело 2ч 15м Z2 + 3×10м под АэП ~155-160 НА РАВНИНЕ",      "вело",    2.25, "",                                              "",       0),
    ("2026-05-21", 4, "Чт", "бег 1ч Z1 + растяжка",                                    "бег",     1.0,  "",                                              "",       0),
    ("2026-05-22", 4, "Пт", "вело 1ч 30м Z2 + 5×30 сек В ПОДЪЁМ",                       "вело",    1.5,  "🤸 СТ ОМВ-ВЕРХ УЛИЦА 1ч 15м (10м разм. бег + 5 кругов ПИК + 10м зам)", "ст-омв-пик",1.25),
    ("2026-05-23", 4, "Сб", "длит. вело 3ч 20м Z1-Z2 (макс субб. мая)",                "вело-длит",3.33, "",                                              "",       0),
    ("2026-05-24", 4, "Вс", "бег 2ч 30м Z2 (рельеф) — длинная мая",                    "бег-длит",2.5,  "",                                              "",       0),

    # W5 (~10.75ч): Разгрузка + переход на бег + контрольный 5К
    ("2026-05-25", 5, "Пн", "отдых ИЛИ прогулка 30м",                                  "отдых",   0,    "",                                              "",       0),
    ("2026-05-26", 5, "Вт", "вело 1ч 15м Z1 (восст)",                                  "вело",    1.25, "🤸 СТ ОМВ ТОНИЗИР. УЛИЦА 45м (5м разм + 2 круга + 5м зам)", "ст-тон", 0.75),
    ("2026-05-27", 5, "Ср", "бег 1ч 30м Z2 (плавный возврат)",                         "бег",     1.5,  "",                                              "",       0),
    ("2026-05-28", 5, "Чт", "вело 1ч 30м Z2",                                          "вело",    1.5,  "",                                              "",       0),
    ("2026-05-29", 5, "Пт", "бег 1ч 15м Z2 + 4×30 сек разгон НА РАВНИНЕ",                "бег",     1.25, "",                                              "",       0),
    ("2026-05-30", 5, "Сб", "вело 3ч Z2 (короткая в тейпере)",                          "вело",    3.0,  "",                                              "",       0),
    ("2026-05-31", 5, "Вс", "🏁 контр. 5К: 30м разм + 5К + 15м зам",                    "контроль",1.5,  "",                                              "",       0),
]

WEEKS_INFO = {
    1: ("W1 хвост", "Втягивающая (1-3.05)"),
    2: ("W2 (4-10.05)", "Вело база 1 + СТАРТ блока СТ ОМВ"),
    3: ("W3 (11-17.05)", "Вело база 2 + СТ ОМВ развивающий"),
    4: ("W4 (18-24.05)", "🔴 НАГРУЗОЧНЫЙ ПИК мезоцикла"),
    5: ("W5 (25-31.05)", "Разгрузка + переход на бег + 🏁 контр. 5К"),
}

THIN = Side(border_style="thin", color="BDBDBD")
MEDIUM = Side(border_style="medium", color="616161")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_HEADER = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)


def fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def style_cell(cell, ttype: str | None, bold: bool = False, size: int = 9):
    if ttype and ttype in TYPE_COLORS and TYPE_COLORS[ttype] != "FFFFFF":
        cell.fill = fill(TYPE_COLORS[ttype])
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    color = "FFFFFF" if ttype in WHITE_ON else "000000"
    cell.font = Font(size=size, bold=bold, color=color)
    cell.border = BORDER_ALL


def build_calendar_sheet(wb: Workbook) -> None:
    """Лист 1: календарь с 2-строчными ячейками (утро/вечер)."""
    ws = wb.create_sheet("Календарь", 0)
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Май 2026 — план по дням (утро / вечер)"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:I1")

    ws["A2"] = "Цель: вело-фокус + 3-нед блок СТ ОМВ + 🏁 контрольный 5К на исход"
    ws["A2"].font = Font(size=10, italic=True, color="616161")
    ws.merge_cells("A2:I2")

    # Шапка
    headers = ["Неделя", "Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс", "Часы"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill("455A64")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER_HEADER

    by_week: dict[int, dict[str, dict]] = {}
    for d in PLAN:
        by_week.setdefault(d[1], {})[d[2]] = {
            "date": d[0],
            "m_text": d[3], "m_type": d[4], "m_h": d[5],
            "e_text": d[6], "e_type": d[7], "e_h": d[8],
        }

    day_order = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]

    row = 5
    for week_num in sorted(by_week.keys()):
        week_label, week_desc = WEEKS_INFO[week_num]
        days_data = by_week[week_num]

        # Левая колонка: 3 строки (метка, утро-маркер, вечер-маркер)
        # Используем 3 строки на неделю: дата, утро, вечер
        date_row = row
        morning_row = row + 1
        evening_row = row + 2

        # Левая ячейка — описание недели — мерж по 3 строкам
        c = ws.cell(row=date_row, column=1, value=f"{week_label}\n{week_desc}")
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.fill = fill("CFD8DC")
        c.border = BORDER_ALL
        ws.merge_cells(start_row=date_row, start_column=1, end_row=evening_row, end_column=1)

        week_total = 0.0
        for col_offset, day_name in enumerate(day_order):
            col = 2 + col_offset
            data = days_data.get(day_name)

            # Строка 1: дата + день недели
            date_short = ""
            if data:
                date_short = f"{day_name} {data['date'][8:10]}.{data['date'][5:7]}"
            dc = ws.cell(row=date_row, column=col, value=date_short)
            dc.font = Font(bold=True, size=9, color="616161")
            dc.alignment = Alignment(horizontal="center", vertical="center")
            dc.fill = fill("F5F5F5")
            dc.border = BORDER_ALL

            if not data:
                ws.cell(row=morning_row, column=col).border = BORDER_ALL
                ws.cell(row=evening_row, column=col).border = BORDER_ALL
                continue

            # Строка 2: УТРО
            m_text = data["m_text"]
            m_type = data["m_type"]
            m_h = data["m_h"]
            if m_text:
                mc = ws.cell(row=morning_row, column=col, value=f"☀ {m_text}")
                style_cell(mc, m_type)
            else:
                mc = ws.cell(row=morning_row, column=col, value="—")
                style_cell(mc, "отдых")

            # Строка 3: ВЕЧЕР
            e_text = data["e_text"]
            e_type = data["e_type"]
            e_h = data["e_h"]
            if e_text:
                ec = ws.cell(row=evening_row, column=col, value=f"🌙 {e_text}")
                style_cell(ec, e_type)
            else:
                ec = ws.cell(row=evening_row, column=col, value="—")
                style_cell(ec, "отдых")
                ec.font = Font(size=8, color="9E9E9E")

            week_total += m_h + e_h

        # Часы недели (мерж 3 строк)
        ch = ws.cell(row=date_row, column=9, value=round(week_total, 1))
        ch.font = Font(bold=True, size=12)
        ch.alignment = Alignment(horizontal="center", vertical="center")
        ch.fill = fill("FFF9C4")
        ch.border = BORDER_ALL
        ws.merge_cells(start_row=date_row, start_column=9, end_row=evening_row, end_column=9)

        ws.row_dimensions[date_row].height = 18
        ws.row_dimensions[morning_row].height = 55
        ws.row_dimensions[evening_row].height = 55
        row += 3

    # Итог
    total_hours = sum(d[5] + d[8] for d in PLAN)
    tc = ws.cell(row=row, column=1, value="ИТОГО ЗА МАЙ")
    tc.font = Font(bold=True, color="FFFFFF")
    tc.fill = fill("455A64")
    tc.border = BORDER_HEADER
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)

    th = ws.cell(row=row, column=9, value=round(total_hours, 1))
    th.font = Font(bold=True, size=14, color="FFFFFF")
    th.fill = fill("455A64")
    th.alignment = Alignment(horizontal="center", vertical="center")
    th.border = BORDER_HEADER
    ws.row_dimensions[row].height = 25

    # Ширина колонок
    ws.column_dimensions["A"].width = 22
    for col_letter in "BCDEFGH":
        ws.column_dimensions[col_letter].width = 32
    ws.column_dimensions["I"].width = 8

    # Легенда
    legend_row = row + 3
    ws.cell(row=legend_row, column=1, value="Легенда (цвета):").font = Font(bold=True, size=11)
    legend = [
        ("вело", "Вело база (Z1-Z2, ровный темп)"),
        ("вело-длит", "Длительная вело (3-5 ч на СБ)"),
        ("бег", "Бег Z1-Z2 (1ч)"),
        ("бег-длит", "Длительный бег (90-120 мин)"),
        ("ст-омв", "СТ ОМВ развивающая (3-4 круга, см. лист «Силовая»)"),
        ("ст-омв-пик", "СТ ОМВ ПИК блока (5 кругов, W4)"),
        ("ст-тон", "СТ ОМВ тонизирующая (1-2 круга, лёгко)"),
        ("контроль", "Контрольный старт / тест"),
        ("отдых", "Отдых / нет тренировки"),
    ]
    for i, (key, desc) in enumerate(legend):
        c = ws.cell(row=legend_row + 1 + i, column=1)
        c.fill = fill(TYPE_COLORS[key])
        c.value = key
        c.font = Font(bold=True, size=9, color="FFFFFF" if key in WHITE_ON else "000000")
        c.border = BORDER_ALL
        ws.cell(row=legend_row + 1 + i, column=2, value=desc).font = Font(size=10)


def build_strength_sheet(wb: Workbook) -> None:
    """Лист 2: СТ ОМВ — методика, разделение зал/улица, упражнения."""
    ws = wb.create_sheet("Силовая ОМВ")
    ws.sheet_view.showGridLines = False

    row = 1
    ws.cell(row=row, column=1, value="СТ ОМВ — методика и упражнения (зал ВТ + улица ПТ)").font = Font(size=14, bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4); row += 1

    ws.cell(row=row, column=1, value="Статодинамика, по Шишкиной/Мякинченко/Платонову — гипертрофия медленных (окислительных) мышечных волокон").font = Font(size=10, italic=True, color="616161")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4); row += 2

    # Расписание силовой
    ws.cell(row=row, column=1, value="📅 Расписание").font = Font(size=12, bold=True)
    row += 1
    sched_headers = ["День", "Время", "Место", "Группа мышц", "Замечания"]
    for col, h in enumerate(sched_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill("455A64")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER_HEADER
    row += 1
    schedule = [
        ("ВТ", "🌙 2-я половина дня (после работы)", "🏋 ЗАЛ", "🦵 НИЗ + кор",
         "Со штангой / гантелями / тренажёрами. Главный силовой день блока."),
        ("ПТ", "🌙 после работы (можно днём)", "🤸 УЛИЦА (спортплощадка)", "💪 ВЕРХ + кор",
         "Турник / брусья / отжимания. Calisthenics, без оборудования."),
    ]
    for s in schedule:
        for col, val in enumerate(s, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = BORDER_ALL
            c.font = Font(size=10)
        ws.row_dimensions[row].height = 35
        row += 1
    row += 1

    # 🔥 РАЗМИНКА — отдельная секция
    ws.cell(row=row, column=1, value="🔥 Разминка перед СТ ОМВ (обязательно!)").font = Font(size=12, bold=True, color="C62828")
    row += 1
    ws.cell(row=row, column=1, value="Разминка ВКЛЮЧЕНА в указанную в плане длительность СТ-сессии (10м разм + работа + 5-10м заминка).").font = Font(size=10, italic=True, color="616161")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 2

    warmup_headers = ["Где", "Этап", "Длит.", "Что делать"]
    for col, h in enumerate(warmup_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = fill("455A64")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER_HEADER
    row += 1
    warmup_rows = [
        ("🏋 ЗАЛ\n(ВТ)", "1. Кардио-разогрев", "10 мин",
         "Велотренажёр Z1 (HR 110-130) ИЛИ беговая дорожка лёгкий бег. Цель — поднять пульс, разогреть кровь, активировать мышцы."),
        ("🏋 ЗАЛ\n(ВТ)", "2. Динамическая растяжка", "5 мин",
         "Махи ногами / выпады с ротацией / круговые движения тазом, плечами, голеностопом. БЕЗ статических удержаний (это после, в заминке)."),
        ("🏋 ЗАЛ\n(ВТ)", "3. Подводящий подход", "перед каждым новым упр. с весом",
         "Перед приседами/становой/жимом — 1-2 подхода с лёгким весом (40-50% рабочего) на 5-8 повторений в обычном темпе. Без отказа. Активирует ЦНС."),
        ("🤸 УЛИЦА\n(ПТ)", "1. Лёгкий бег / трусцой", "8-10 мин",
         "Лёгкий бег вокруг площадки HR 110-130. Если холодно — длиннее (до 15 мин)."),
        ("🤸 УЛИЦА\n(ПТ)", "2. Динамическая растяжка", "5 мин",
         "Махи руками, ротация корпуса, круговые движения плечами, активная растяжка плеч/широчайших."),
        ("🤸 УЛИЦА\n(ПТ)", "3. Подготовка к турнику", "1-2 мин",
         "Вис на турнике 10-15 сек × 2 раза. 5-7 негативных подтягиваний (медленный спуск) — активирует спину/руки."),
        ("ОБА", "🧊 Заминка (после работы)", "5-10 мин",
         "Статическая растяжка работавших мышц (по 20-30 сек на каждую группу). 5 мин лёгкого вело/ходьбы для снижения пульса."),
    ]
    for r in warmup_rows:
        for col, val in enumerate(r, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border = BORDER_ALL
            c.font = Font(size=10)
            if "Заминка" in r[1]:
                c.fill = fill("E1F5FE")
        ws.row_dimensions[row].height = 60
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="📌 Итого разминка + заминка: ~15-20 мин на ВТ (зал), ~15-20 мин на ПТ (улица). Это часть указанной в плане длительности СТ.").font = Font(size=10, italic=True, color="C62828", bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 2

    # Принципы
    ws.cell(row=row, column=1, value="Общие принципы (одинаковые для зала и улицы)").font = Font(size=12, bold=True)
    row += 1
    principles = [
        "• Цель: гипертрофия медленных (окислительных) мышечных волокон — база для аэробной выносливости",
        "• Метод: статодинамика — медленные движения БЕЗ полного разгибания и БЕЗ расслабления",
        "• Интенсивность: 40-60% МПС (средне-лёгкие веса/уровень нагрузки, ощущение жжения, не «надорвать»)",
        "• 1 подход: 30-40 секунд непрерывной работы ДО ОТКАЗА (когда не можешь продолжать с правильной техникой)",
        "• 1 серия = 3 подхода через 25 сек отдыха внутри серии",
        "• Между сериями (упражнениями/кругами) 5-10 мин активного восстановления (ходьба, потягивания)",
        "• Антигликолитика: НЕ закислять выше 4 ммоль/л → длинные паузы между сериями",
        "• Развивающие тренировки 2 раза/нед с минимум 60 ч между ними (поэтому ВТ + ПТ)",
        "• Период полужизни белков НМА = 10-12 дней → 3-нед блок без пропусков, иначе откат",
    ]
    for p in principles:
        c = ws.cell(row=row, column=1, value=p)
        c.font = Font(size=10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
    row += 1

    # Структура по неделям
    ws.cell(row=row, column=1, value="Структура тренировки по неделям").font = Font(size=12, bold=True)
    row += 1
    structure_headers = ["Неделя", "Кол-во кругов", "Подходы 30-40 сек / отдых 25 сек", "Между кругами", "Длит."]
    for col, h in enumerate(structure_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill("455A64")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER_HEADER
    row += 1
    structure = [
        ("W2 — Втягивающая", "3 круга", "3 подхода (30 сек / 25 сек)", "5-7 мин", "60 мин"),
        ("W3 — Развивающая", "4 круга", "3 подхода (30-40 сек / 25 сек)", "6-8 мин", "75 мин"),
        ("W4 — ПИК блока", "5 кругов", "3 подхода (30-40 сек / 25 сек)", "7-10 мин", "75-90 мин"),
        ("W5 — Тонизирующая", "1-2 круга", "1-2 подхода (30 сек / 25 сек) лёгко", "5 мин", "30-45 мин"),
    ]
    for s in structure:
        for col, val in enumerate(s, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = BORDER_ALL
            c.font = Font(size=10)
            if "ПИК" in s[0]:
                c.fill = fill("FFE0B2")
        row += 1
    row += 1

    # 🏋 ЗАЛ — НИЗ (ВТ)
    ws.cell(row=row, column=1, value="🏋 ВТ ЗАЛ — НИЗ + кор (со штангой / гантелями / тренажёрами)").font = Font(size=12, bold=True, color="1976D2")
    row += 1
    headers = ["№", "Упражнение", "Описание (статодинамика)", "Замечания"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill("455A64")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER_HEADER
    row += 1
    ex_gym = [
        (1, "Приседы со штангой",
         "На спине / фронтальные. Темп 3 сек вниз / 2 сек вверх. БЕЗ разгибания коленей до конца, БЕЗ касания/расслабления внизу.",
         "Вес 40-60% от 1ПМ. Глубина комфортная (угол колена ≥90°). Главное упражнение блока."),
        (2, "Жим ногами (тренажёр) ИЛИ выпады с гантелями",
         "На тренажёре: 3/2 сек, БЕЗ полного разгибания. Выпады: поочерёдно, шаг назад/вперёд.",
         "Если выпады — по 30-40 сек на каждую ногу (т.е. 1 подход = 60-80 сек)."),
        (3, "Становая тяга или румынская тяга",
         "Гриф/гантели. Спина прямая, медленный темп. На РТ останавливаемся выше колена, не вниз до пола.",
         "Включает заднюю поверхность бедра + ягодицы. 40-50% 1ПМ. Без разгибания корпуса полностью."),
        (4, "Икры на тренажёре (или со штангой стоя)",
         "Подъём на носки на машине Смита/в станке для икр. Медленный темп 2/3 сек. БЕЗ касания платформы пятками внизу.",
         "Можно с гантелями в руках на ступеньке если без зала. Сильное жжение."),
        (5, "Кор: планка с диском / V-up / скручивания на скамье с весом",
         "Кор-блок 30-40 сек: статика-планка с диском на пояснице ИЛИ V-up с гантелью.",
         "Чередовать варианты по подходам."),
    ]
    for ex in ex_gym:
        for col, val in enumerate(ex, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border = BORDER_ALL
            c.font = Font(size=10)
        ws.row_dimensions[row].height = 60
        row += 1
    row += 1

    # 🤸 УЛИЦА — ВЕРХ (ПТ)
    ws.cell(row=row, column=1, value="🤸 ПТ УЛИЦА — ВЕРХ + кор (турник / брусья / отжимания / собств.вес)").font = Font(size=12, bold=True, color="2E7D32")
    row += 1
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill("455A64")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER_HEADER
    row += 1
    ex_street = [
        (1, "Подтягивания на турнике",
         "Прямой/обратный/нейтральный хват. Темп 3 сек вверх / 3 сек вниз. БЕЗ полного выпрямления рук внизу, БЕЗ касания подбородком/грудью перекладины (выйти на 80%).",
         "Если 30-40 сек подтягиваться полностью не получается — делай частичные (только середина амплитуды) или негатив (медленный спуск). Прогрессия: добавлять вес в рюкзак."),
        (2, "Отжимания на брусьях",
         "Медленный темп 3/2 сек. БЕЗ полного выпрямления локтей сверху. Глубина — плечо чуть ниже параллели локтя.",
         "Если 30-40 сек на брусьях тяжело — частичные (только верх амплитуды) или замени на отжимания узким хватом."),
        (3, "Отжимания от пола",
         "Классические или с упором на брусья. 3/2 сек. БЕЗ полного выпрямления рук, БЕЗ касания груди пола.",
         "Прогрессия: руки на возвышении (для груди) или ноги на возвышении (для плеч). С рюкзаком — для прогрессии."),
        (4, "Австралийские подтягивания (низкая перекладина)",
         "Низкая перекладина (на уровне пояса). Тело прямое, тянемся грудью к перекладине. 3/2 сек, БЕЗ полного выпрямления внизу.",
         "Альтернатива/дополнение к классическим подтягиваниям. Прорабатывает спину горизонтально."),
        (5, "Кор на турнике / на полу",
         "Подъём ног на турнике (висом) до уровня живота — 30-40 сек. ИЛИ планка / V-up на полу.",
         "На турнике сложнее (плюс работа хвата). Чередовать варианты."),
    ]
    for ex in ex_street:
        for col, val in enumerate(ex, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border = BORDER_ALL
            c.font = Font(size=10)
        ws.row_dimensions[row].height = 70
        row += 1
    row += 1

    # Поминутный план
    ws.cell(row=row, column=1, value="Поминутный план тренировки (пример развив. W3, 75 мин)").font = Font(size=12, bold=True)
    row += 1
    detail_headers = ["Время", "Что"]
    for col, h in enumerate(detail_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill("455A64")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER_HEADER
    row += 1
    timing = [
        ("0-10'", "Разминка: 5 мин кардио (вело/беговая дорожка) + 5 мин динам. растяжка / мобилизация суставов"),
        ("10-13'", "Круг 1, упр. 1 — 3 подхода (30\" работа / 25\" отдых)"),
        ("13-16'", "Круг 1, упр. 2 — 3 подхода"),
        ("16-19'", "Круг 1, упр. 3 — 3 подхода"),
        ("19-22'", "Круг 1, упр. 4 — 3 подхода"),
        ("22-25'", "Круг 1, упр. 5 — 3 подхода (кор)"),
        ("25-31'", "Между кругами 6 мин: ходьба / потягивания / лёгкое кардио"),
        ("31-46'", "КРУГ 2 (упр. 1-5, та же схема)"),
        ("46-52'", "Между кругами 6 мин"),
        ("52-67'", "КРУГ 3 (упр. 1-5)"),
        ("67-73'", "Между кругами 6 мин"),
        ("73-75'", "Заминка: растяжка 2 мин (полную сделай позже)"),
    ]
    for t in timing:
        for col, val in enumerate(t, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = BORDER_ALL
            c.font = Font(size=10)
        row += 1

    # Ширина
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 45


def build_glossary_sheet(wb: Workbook) -> None:
    """Лист: справка по сокращениям."""
    ws = wb.create_sheet("Сокращения")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Справка по сокращениям и терминам"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:D1")

    ws["A2"] = "Все аббревиатуры используемые в плане и листах файла"
    ws["A2"].font = Font(size=10, italic=True, color="616161")
    ws.merge_cells("A2:D2")

    sections = [
        ("Мышечные волокна", [
            ("ОМВ", "Окислительные (медленные) мышечные волокна",
             "Тип I по Henneman. Главные при работе ниже ПАНО. Богаты митохондриями. Плохо гипертрофируются от обычных силовых, только статодинамика/квазиизометрия. ОСНОВНАЯ ЦЕЛЬ нашего силового блока в мае-июне."),
            ("БМВ", "Быстрые мышечные волокна",
             "Тип II. Гипертрофируются от обычной силовой/прыжков. Включаются при работе выше ПАНО или при максимальных усилиях."),
            ("БоМВ", "Быстрые окислительные МВ",
             "Тип IIA. Промежуточный тип — могут быть и быстрыми, и аэробными. Включаются на ПАНО."),
            ("БгМВ", "Быстрые гликолитические МВ",
             "Тип IIB/IIX. Чисто анаэробные, спринтерские. В лыжных гонках играют роль только в финишных спринтах."),
            ("ППС", "Площадь поперечного сечения МВ",
             "Чем больше ППС — тем сильнее волокно. Но обратная зависимость с окислительной способностью (см. парадокс Хиксона)."),
        ]),
        ("Силовые показатели", [
            ("1ПМ (1RM)", "Повторный максимум (One Rep Max)",
             "Максимальный вес, с которым можешь выполнить ОДНО повторение упражнения. Определяется тестом или по таблицам пересчёта (например, 5×80 кг ≈ 1ПМ 92 кг)."),
            ("МПС", "Максимальная произвольная сила",
             "Аналог 1ПМ, иногда подразумевает изометрическое максимальное усилие. В контексте плана — синоним 1ПМ."),
            ("BW", "Bodyweight, собственный вес",
             "Используется в калистенике (подтягивания, отжимания) — нагрузка = твой вес."),
            ("TUT", "Time Under Tension, время под нагрузкой",
             "В статодинамике 30-40 сек на подход — целевое TUT для гипертрофии ОМВ."),
        ]),
        ("Пороги и зоны", [
            ("АэП / LT1", "Аэробный порог (~2 ммоль/л лактата)",
             "Точка где лактат начинает расти. Ниже — чистая аэробика (LIT). Твой АэП = 160 уд/мин."),
            ("АнП / ПАНО / LT2", "Анаэробный порог = ВЛ (вентиляционный лимит)",
             "Точка где лактат растёт неконтролируемо (~4 ммоль). Граница sweet spot и HIT. Твой АнП = 178 уд/мин."),
            ("LT1 / LT2", "Lactate Threshold 1 и 2",
             "Английские эквиваленты АэП и АнП. Используются как взаимозаменяемые с русскими."),
            ("МПК / VO2max", "Максимальное потребление кислорода",
             "Мл O₂ на кг массы тела за минуту. У тебя расч. ~62-65 (по vПАНО). Главный аэробный потолок."),
            ("Z1...Z5", "Зоны интенсивности",
             "5-зонная модель. Z1<140 / Z2 140-161 / Z3 162-178 / Z4 179-185 / Z5 186+ — твоя калибровка по лактат-тесту."),
            ("LIT / MIT / HIT", "Low / Moderate / High Intensity Training",
             "3-зонная модель Сейлера. LIT = <АэП (Z1+Z2). MIT = АэП-АнП (Z3). HIT = >АнП (Z4+Z5)."),
        ]),
        ("Сердце и восстановление", [
            ("ЧСС / HR", "Частота сердечных сокращений",
             "Heart Rate. Измеряется уд/мин."),
            ("HRmax", "Максимальная ЧСС",
             "Максимальный пульс который ты способен достичь. У тебя 187."),
            ("LTHR", "Lactate Threshold Heart Rate",
             "ЧСС на анаэробном пороге. Garmin использует это поле для пересчёта зон. Твой LTHR = 178."),
            ("RHR", "Resting Heart Rate, ЧСС покоя",
             "Утренний пульс лежа. Маркер восстановления — у тебя база ~46-48."),
            ("HRV", "Heart Rate Variability, вариабельность ЧСС",
             "Отражает баланс симпатической/парасимпатической НС. Маркер готовности к нагрузке. Чем выше — тем лучше."),
            ("ССС", "Сердечно-сосудистая система",
             "Сердце + сосуды. У тебя «сниженный ударный объём сердца» (Федосеев 2023) — морфологическое."),
        ]),
        ("Системы организма", [
            ("НМА", "Нервно-мышечный аппарат",
             "Период полужизни белков НМА = 10-12 дней — ориентир длительности силового мезоцикла."),
            ("МАМ", "Максимальная алактатная мощность",
             "Способность мышцы генерировать пиковую мощность за 5-10 секунд без накопления лактата."),
        ]),
        ("Тренировки и периодизация", [
            ("СТ", "Силовая тренировка",
             "В плане часто пишется «СТ ОМВ», «СТ НСВИ» и т.п. — указание режима/направленности."),
            ("ОФП", "Общая физическая подготовка",
             "Базовые упражнения для общей формы (без специфики вида спорта)."),
            ("СФП", "Специальная физическая подготовка",
             "Упражнения максимально близкие к соревновательной локомоции (имитация лыжного хода и т.п.)."),
            ("ЦВС", "Циклические виды спорта",
             "Бег, лыжи, плавание, гребля, велоспорт — все где локомоция повторяется."),
            ("СК", "Суперкомпенсация",
             "Период (7-14 дней) после нагрузочного блока, когда организм восстанавливается + увеличивает функциональность выше исходного уровня."),
        ]),
        ("Режимы силовой/циклической работы (Доклад Перм)", [
            ("НСНИ", "Низкоскоростной низкоинтенсивный",
             "Низкая скорость движений + низкая нагрузка. Например, длительная аэробика на лёгком темпе."),
            ("НСВИ", "Низкоскоростной высокоинтенсивный",
             "Низкая скорость + высокая нагрузка. Это макс.сила (приседы со штангой 85-95% 1ПМ, 2-4 повт). В плане: июль и поддерж. летом-осенью."),
            ("НССИ", "Низкоскоростной среднеинтенсивный",
             "Низкая скорость + средняя нагрузка. = СТАТОДИНАМИКА для гипертрофии ОМВ. ТО ЧТО МЫ ДЕЛАЕМ В МАЕ."),
            ("ВСВИ", "Высокоскоростной высокоинтенсивный",
             "Высокая скорость + высокая нагрузка. = взрывная/плиометрика/спринт в гору. Гипертрофия БМВ."),
            ("АСМ", "Аэробно-силовой метод",
             "Циклика с дополнительным сопротивлением (роллеры, бег по песку, в гору). Активирует ММВ при аэробном режиме."),
            ("АММ", "Аэробно-мощностной метод",
             "Высокоинтенсивная циклика на ПАНО или выше — для повышения мощности."),
        ]),
        ("Тейпер и подводка", [
            ("Тейпер (taper)", "Подводка к старту со снижением объёма",
             "Объём -30-50% за 1-2 нед до главного старта при сохранении интенсивности. Цель — сохранить форму, разгрузить организм."),
            ("A-старт / B-старт", "Главный / тренировочный старт",
             "A — главный (полный тейпер 13 дней). B — по дороге, тренировочный (короткий тейпер 5-7 дней, не разваливать блок)."),
            ("ЭКМ / ГС", "Этап Кубка Мира / Главный старт",
             "Из Доклада Перм — терминология сборной."),
        ]),
        ("Мониторинг", [
            ("TE", "Training Effect (Garmin)",
             "Оценка тренировочного воздействия 0-5: 0-1 нет / 1-2 поддержание / 2-3 улучшение / 3-4 высокое / 4-5 экстремальное."),
            ("Body Battery", "Шкала энергии Garmin",
             "0-100 на основе HRV+стресс+сон. Маркер готовности к нагрузке."),
        ]),
    ]

    row = 4
    for section_name, items in sections:
        # Заголовок секции
        c = ws.cell(row=row, column=1, value=section_name)
        c.font = Font(bold=True, size=12, color="FFFFFF")
        c.fill = fill("455A64")
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 22
        row += 1

        # Шапка таблицы
        for col, h in enumerate(["Сокр.", "Расшифровка", "Описание"], 1):
            cc = ws.cell(row=row, column=col, value=h)
            cc.font = Font(bold=True, size=10, color="FFFFFF")
            cc.fill = fill("78909C")
            cc.alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        for abbr, full, desc in items:
            ws.cell(row=row, column=1, value=abbr).font = Font(bold=True, size=10)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="top")
            ws.cell(row=row, column=2, value=full).font = Font(size=10)
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.cell(row=row, column=3, value=desc).font = Font(size=10, color="424242")
            ws.cell(row=row, column=3).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = BORDER_ALL
            ws.row_dimensions[row].height = 35
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 75


def build_weights_sheet(wb: Workbook) -> None:
    """Лист: расчёт примерных весов отягощения от 81-82 кг."""
    ws = wb.create_sheet("Веса (81-82 кг)")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Расчёт примерных весов отягощения"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:F1")

    ws["A2"] = "Атлет: 81-82 кг, тренированный любитель/полупрофи. Без специфической силовой подготовки последние годы."
    ws["A2"].font = Font(size=10, italic=True, color="616161")
    ws.merge_cells("A2:F2")

    row = 4
    # Объяснение
    ws.cell(row=row, column=1, value="🟡 ВАЖНО: цифры ниже — ОРИЕНТИР, не догма.").font = Font(bold=True, size=11, color="C62828")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    ws.cell(row=row, column=1, value="В статодинамике главный критерий — ОТКАЗ на 30-40 секунде. Калибровка по факту:").font = Font(size=10)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    rules = [
        "  • Если дошёл до отказа РОВНО на 30-35 сек — вес ПРАВИЛЬНЫЙ",
        "  • Если еле дотянул до 20-25 сек — вес ВЕЛИК → -15-20%",
        "  • Если можешь продолжать после 40-45 сек — вес МАЛ → +15-20%",
        "  • На 1-м подходе серии — выбери целевой вес. Если 2-й/3-й подходы тяжелее (норма) — вес тот же, не снижай",
    ]
    for r in rules:
        ws.cell(row=row, column=1, value=r).font = Font(size=10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
    row += 1

    # Оценка 1ПМ для тренированного атлета 81-82 кг
    ws.cell(row=row, column=1, value="📊 Оценочный 1ПМ (One Rep Max) для тренированного атлета 81 кг").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="(Если у тебя есть свой замер 1ПМ — лучше используй его и пересчитай 40-50%)").font = Font(size=10, italic=True, color="616161")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 2

    # ЗАЛ
    ws.cell(row=row, column=1, value="🏋 ЗАЛ — Вторник (НИЗ)").font = Font(bold=True, size=12, color="1976D2")
    row += 1

    headers_gym = ["Упражнение", "1ПМ оценка", "40% (W2 втягив.)", "50% (W3 развив.)", "55-60% (W4 ПИК)", "Замечания"]
    for col, h in enumerate(headers_gym, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = fill("455A64")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER_HEADER
    ws.row_dimensions[row].height = 35
    row += 1

    weights_gym = [
        ("Приседы со штангой", "100-120 кг", "40-48 кг", "50-60 кг", "55-72 кг",
         "Главное упражнение. Старт на W2 = 40-45 кг (~½ BW). Темп 3 сек вниз / 2 вверх."),
        ("Становая / румынская тяга", "120-140 кг", "48-56 кг", "60-70 кг", "66-84 кг",
         "Старт W2 = 50 кг. РТ — без полного выпрямления + остановка выше колена."),
        ("Жим ногами (тренажёр)", "180-220 кг", "72-88 кг", "90-110 кг", "99-132 кг",
         "На тренажёре можно больше, чем приседы. Без полного разгибания коленей сверху."),
        ("Выпады с гантелями (если без жима)", "—", "10-12 кг (×2)", "14-16 кг (×2)", "16-20 кг (×2)",
         "Гантели в каждую руку. По 30-40 сек на ногу = 1 подход."),
        ("Икры на тренажёре", "100-130 кг", "40-52 кг", "50-65 кг", "55-78 кг",
         "Можно близко к 1ПМ, икры быстро адаптируются. Полная амплитуда низкоконтролируемо."),
        ("Кор: V-up с гантелью / планка с диском", "—", "8-10 кг", "10-12 кг", "12-15 кг",
         "Лёгкое отягощение, главное — TUT 30-40 сек."),
    ]
    for w in weights_gym:
        for col, val in enumerate(w, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border = BORDER_ALL
            c.font = Font(size=10)
        ws.row_dimensions[row].height = 50
        row += 1
    row += 1

    # УЛИЦА
    ws.cell(row=row, column=1, value="🤸 УЛИЦА — Пятница (ВЕРХ, calisthenics)").font = Font(bold=True, size=12, color="2E7D32")
    row += 1

    headers_street = ["Упражнение", "База", "W2 втягив.", "W3 развив.", "W4 ПИК", "Замечания"]
    for col, h in enumerate(headers_street, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = fill("455A64")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER_HEADER
    ws.row_dimensions[row].height = 35
    row += 1

    weights_street = [
        ("Подтягивания на турнике", "BW (81 кг)",
         "BW, чистые подтягивания 30 сек ИЛИ негативы (3 сек спуск)",
         "BW, если в чистую делаешь 30-35 сек до отказа",
         "BW + 5 кг рюкзак, если на чистом легко",
         "Если 30 сек чистых не получается — частичные (только верх амплитуды) или негативы. Прогрессия — добавлять 2.5-5 кг в рюкзак."),
        ("Отжимания на брусьях", "BW (81 кг)",
         "BW (если 30 сек тяжело — частичные)",
         "BW",
         "BW + 5-10 кг рюкзак",
         "Без полного выпрямления локтей сверху. Если на брусьях невозможно — отжимания узким хватом от пола (BW)."),
        ("Отжимания от пола", "BW (~50 кг нагрузка на руки)",
         "BW классические",
         "BW + руки на возвышении 20-30 см ИЛИ ноги на возвышении",
         "BW + рюкзак 5-10 кг ИЛИ ноги на скамье",
         "Темп 3/2 сек, без полного выпрямления, без касания пола."),
        ("Австралийские подтягивания", "BW (~60-70 кг нагрузка)",
         "BW, перекладина на уровне пояса",
         "BW, перекладина чуть ниже + ноги на возвышении (повышение нагрузки)",
         "BW + рюкзак 5 кг ИЛИ ноги на возвышении 30 см",
         "Альтернатива/дополнение к классическим подтягиваниям. Хорошо для горизонтальной тяги."),
        ("Кор на турнике (подъём ног)", "BW",
         "Подъём согнутых ног 30 сек",
         "Подъём прямых ног до уровня живота 30-35 сек",
         "Подъём прямых ног до перекладины 30-35 сек",
         "На турнике сложно — много работы хвата. Альтернатива на полу: V-up / планка с подъёмом ног."),
    ]
    for w in weights_street:
        for col, val in enumerate(w, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border = BORDER_ALL
            c.font = Font(size=10)
        ws.row_dimensions[row].height = 65
        row += 1
    row += 1

    # Тест 1ПМ
    ws.cell(row=row, column=1, value="🧪 Если хочешь точно: тест 1ПМ").font = Font(bold=True, size=12)
    row += 1
    test_lines = [
        "Опционально, до W2: один раз посчитай 1ПМ для приседов и становой по такой схеме:",
        "  1. Разминка 10 мин кардио + динам. растяжка",
        "  2. Несколько прогревочных подходов с возрастающим весом (например 40 кг ×8, 60 кг ×5, 80 кг ×3, 90 кг ×2)",
        "  3. Затем подбираешь вес который можешь поднять 4-6 раз — и пересчитываешь по формуле Эпли:",
        "     1ПМ ≈ Вес × (1 + 0.0333 × Повторения)",
        "     Пример: 90 кг ×4 → 1ПМ ≈ 90 × (1 + 0.0333 × 4) = 90 × 1.133 = 102 кг",
        "  4. Отдых 3-5 мин между подходами. НЕ доходи до отказа на 1ПМ-тесте — травмоопасно",
        "  5. Затем 40-60% от полученного 1ПМ — твой рабочий вес для статодинамики",
        "",
        "⚠️ Для статодинамики (TUT 30-40 сек) фактический рабочий вес обычно ~30-50% 1ПМ — то есть НИЖЕ обычной силовой.",
    ]
    for line in test_lines:
        ws.cell(row=row, column=1, value=line).font = Font(size=10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

    # Ширина
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 50


def build_intervals_sheet(wb: Workbook) -> None:
    """Лист 3: ускорения и темповые с указанием рельефа."""
    ws = wb.create_sheet("Ускорения и темп")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Ускорения / темповые / повторки — рельеф"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:E1")

    ws["A2"] = "Принцип: для каждой интервальной/темповой сессии явно указан рельеф (подъём / равнина / трек / рельеф). Используется и в будущих планах (июнь+)."
    ws["A2"].font = Font(size=10, italic=True, color="616161")
    ws.merge_cells("A2:E2")

    headers = ["Дата", "День", "W", "Что (отрезки)", "Рельеф", "Зачем"]
    row = 4
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill("455A64")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER_HEADER

    intervals = [
        # (дата, день, W, что, рельеф, зачем)
        ("01.05", "Пт", "W1", '4×30" разгон в конце вело-сессии',
         "🟢 НА РАВНИНЕ (ровный участок шоссе/трек)",
         "Нейропатерны, активация ног после длительной разминки. Без закисления."),
        ("08.05", "Пт", "W2", '4×30" в подъём',
         "🔴 В ПОДЪЁМ (5-8°, средняя крутизна)",
         "Активация БМВ + нагрузка перед СТ ОМВ-верх вечером. Восстановление до HR 110-115 между."),
        ("13.05", "Ср", "W3", '3-4×20" ускорения',
         "🔴 В КОРОТКИЙ ПОДЪЁМ (10-15 сек подъёма достаточно)",
         "Поддержание скоростно-силового тонуса, без полного восстановления (вело Z2 между)."),
        ("20.05", "Ср", "W4", "2×10' под АэП ~155-160",
         "🟢 НА РАВНИНЕ (ровный темп)",
         "Темповая в Z2-верх — поддержание АэП-зоны во время нагрузочного пика."),
        ("22.05", "Пт", "W4", '5×30" в подъём',
         "🔴 В ПОДЪЁМ (5-8°)",
         "Пик блока ОМВ — последняя силовая активация перед СТ-вечером."),
        ("29.05", "Пт", "W5", '4×30" разгон',
         "🟢 НА РАВНИНЕ (ровный темп)",
         "Тейпер — поддержание скорости/нейропатернов перед контрольным 5К."),
        ("31.05", "Вс", "W5", "🏁 контрольный 5К + разм/зам",
         "🟢 НА РАВНИНЕ (трек или ровный шоссе) — для чистого замера темпа",
         "Бенчмарк vПАНО перед началом бегового блока в июне."),
    ]
    row = 5
    for d in intervals:
        for col, val in enumerate(d, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border = BORDER_ALL
            c.font = Font(size=10)
            # Цвет ячейки рельефа
            if col == 5:
                if "ПОДЪЁМ" in str(val):
                    c.fill = fill("FFCDD2")  # красноватый
                elif "РАВНИНЕ" in str(val) or "трек" in str(val).lower():
                    c.fill = fill("C8E6C9")  # зелёный
        ws.row_dimensions[row].height = 50
        row += 1

    row += 2

    # Правила записи рельефа
    ws.cell(row=row, column=1, value="Правила обозначения рельефа").font = Font(bold=True, size=12)
    row += 1
    rules = [
        "🟢 НА РАВНИНЕ — ровный участок шоссе/тропы/трек, угол наклона 0-1%",
        "🔴 В ПОДЪЁМ — наклон 4-10%, средняя крутизна (можно отрабатывать силу отталкивания/мощность)",
        "🔴 В КОРОТКИЙ ПОДЪЁМ — крутой и короткий (10-15 сек), для активации БМВ без накопления лактата",
        "🟡 ПО РЕЛЬЕФУ — холмистая трасса, естественная вариативность (для длительных Z2)",
        "🟢 ТРЕК — стадион 400 м, для контрольных и точных замеров темпа",
    ]
    for r in rules:
        c = ws.cell(row=row, column=1, value=r)
        c.font = Font(size=10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="📌 ВАЖНО: восстановление между отрезками в подъём — до HR 110-115 (а не по таймеру). Это антигликолитика — не закислять выше 4 ммоль/л.").font = Font(size=10, italic=True, color="C62828")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    # Ширина колонок
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 55


def build_linear_sheet(wb: Workbook) -> None:
    """Лист 3: линейный список всех тренировок."""
    ws = wb.create_sheet("Тренировки списком")

    headers = ["Дата", "День", "W", "Часть дня", "Тренировка", "Тип", "Часы"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill("455A64")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER_HEADER

    row = 2
    for d in PLAN:
        date, week, day, m_text, m_type, m_h, e_text, e_type, e_h = d
        for part, text, ttype, hrs in [("утро", m_text, m_type, m_h), ("вечер", e_text, e_type, e_h)]:
            if not text:
                continue
            ws.cell(row=row, column=1, value=date)
            ws.cell(row=row, column=2, value=day)
            ws.cell(row=row, column=3, value=week)
            ws.cell(row=row, column=4, value=part)
            ws.cell(row=row, column=5, value=text)
            ws.cell(row=row, column=6, value=ttype)
            ws.cell(row=row, column=7, value=hrs)
            if ttype in TYPE_COLORS:
                font_color = "FFFFFF" if ttype in WHITE_ON else "000000"
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = fill(TYPE_COLORS[ttype])
                    ws.cell(row=row, column=col).border = BORDER_ALL
                    ws.cell(row=row, column=col).font = Font(size=10, color=font_color)
            row += 1

    widths = [12, 6, 5, 9, 60, 14, 7]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_summary_sheet(wb: Workbook) -> None:
    """Лист 4: сводка."""
    ws = wb.create_sheet("Сводка")

    ws["A1"] = "Сводка по маю"
    ws["A1"].font = Font(size=14, bold=True)

    ws["A3"] = "По неделям"
    ws["A3"].font = Font(bold=True, size=11)
    headers = ["Неделя", "Часы", "Описание"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill("455A64")

    weekly = {}
    for d in PLAN:
        weekly.setdefault(d[1], 0)
        weekly[d[1]] += d[5] + d[8]

    row = 5
    for w in sorted(weekly):
        label, desc = WEEKS_INFO[w]
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=round(weekly[w], 1))
        ws.cell(row=row, column=3, value=desc)
        row += 1
    ws.cell(row=row, column=1, value="ИТОГО").font = Font(bold=True)
    ws.cell(row=row, column=2, value=round(sum(weekly.values()), 1)).font = Font(bold=True, size=12)

    row += 3
    ws.cell(row=row, column=1, value="По типам тренировок").font = Font(bold=True, size=11)
    row += 1
    headers = ["Тип", "Часы", "% от месяца"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill("455A64")

    by_type = {}
    for d in PLAN:
        if d[4]:
            by_type[d[4]] = by_type.get(d[4], 0) + d[5]
        if d[7]:
            by_type[d[7]] = by_type.get(d[7], 0) + d[8]
    total = sum(by_type.values())

    rr = row + 1
    for t, h in sorted(by_type.items(), key=lambda x: -x[1]):
        c = ws.cell(row=rr, column=1, value=t)
        if t in TYPE_COLORS:
            c.fill = fill(TYPE_COLORS[t])
            if t in WHITE_ON:
                c.font = Font(color="FFFFFF", bold=True)
        ws.cell(row=rr, column=2, value=round(h, 1))
        ws.cell(row=rr, column=3, value=f"{round(100 * h / total, 1)}%")
        rr += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 35


def main():
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    build_calendar_sheet(wb)
    build_strength_sheet(wb)
    build_weights_sheet(wb)
    build_intervals_sheet(wb)
    build_glossary_sheet(wb)
    build_linear_sheet(wb)
    build_summary_sheet(wb)

    saved = save_with_fallback(wb, OUT)
    print(f"Сохранено: {saved}")
    if saved != OUT:
        print(f"  ⚠ Основной файл {OUT} занят (открыт в Excel) — закрой его, потом перегенерируй.")
    total = sum(d[5] + d[8] for d in PLAN)
    print(f"Сумма часов: {total:.1f}")


if __name__ == "__main__":
    main()
