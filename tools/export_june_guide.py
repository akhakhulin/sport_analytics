"""Выгружает 24-дневный гайд (19.05 — 12.06) в Excel.

Формат:
- Лист "Календарь" — день за днём, с фазой, планом, рекомендацией, ЧСС-потолком, БАДами
- Лист "Утренний decision tree" — пороги HRV/RHR/BB → действие
- Лист "БАДы по фазам"
- Лист "Тесты и чек-листы"

Выход: plans/corrections_to_12-06-2026.xlsx
"""
from __future__ import annotations

import sys
from datetime import date, timedelta
from pathlib import Path

# Импорт plan_reader
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from bot.plan_reader import for_date  # noqa: E402

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# === Стили ===
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_PHASE_A = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # бледно-жёлтый
FILL_PHASE_B = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")  # бледно-голубой
FILL_PHASE_C = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")  # бледно-зелёный
FILL_RACE = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")     # бледно-красный
FILL_HEADER = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")  # тёмный
FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)
FONT_BOLD = Font(bold=True)
ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


WEEKDAYS_RU = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]


def phase_for_date(d: date) -> tuple[str, str]:
    """Возвращает (имя фазы, заливка-ключ)."""
    if d <= date(2026, 5, 28):
        return ("A: wind-in", "A")
    if d <= date(2026, 6, 6):
        return ("B: раскачка + лактат", "B")
    if d <= date(2026, 6, 11):
        return ("C: подводка", "C")
    if d == date(2026, 6, 12):
        return ("🏁 СТАРТ 2.5К", "RACE")
    return ("после старта", "—")


def recommendation_for_day(d: date, sessions: list, phase_key: str) -> tuple[str, str, str]:
    """Возвращает (рекомендация, HR-потолок, заметки)."""
    if phase_key == "RACE":
        return (
            "🏁 КОНТРОЛЬНАЯ 2.5К на максимум",
            "до HRmax 187",
            "Pre: Цитрулин 6г + Гипоксен 500мг + Стимол. После: восст. 3 дня",
        )

    plans_text = " / ".join(f"{s.part}: {s.text}" for s in sessions) if sessions else "—"

    if phase_key == "A":
        # Wind-in — снижаем интенсивность всех Z2 сессий
        return (
            "По плану xlsx, но без E1+. Длительные без подъёмов в Z3.",
            "Z1≤140 / Z2≤155 (бег), Z2≤150 (вело)",
            "СТ только поддерживающая (2-3 круга), без до-отказа. Жара >22°C → бег→вело.",
        )

    if phase_key == "B":
        # Раскачка
        if d == date(2026, 6, 2):  # Пн — возможный МПК
            return (
                "МПК Семёновский: 4-5×5' / 8' R (E2, лактат 5-6 ммоль/л)",
                "Темп ~3:55-4:00, HR верх Z3-низ Z4",
                "ВАЖНО: только Семёновский формат. МПК+ 15\"/45\" — на июль после ретеста.",
            )
        if d in (date(2026, 6, 4), date(2026, 6, 5)):
            return (
                "Лактат-ретест в лаборатории (запиши заранее)",
                "Тестовый ступенчатый",
                "1 день отдыха до теста, 1 день лёгкий после",
            )
        return (
            "По плану xlsx. E1 разрешён.",
            "Z1-Z2 по плану, Z3 только в МПК-сессии",
            "Длительные с лёгким рельефом ОК (≤10 м/км)",
        )

    if phase_key == "C":
        # Подводка
        plan_C = {
            date(2026, 6, 7): ("Сб: A1 40 мин", "HR ≤140", "Объём -50% от пика"),
            date(2026, 6, 8): ("Вс: A1 30 мин или отдых", "HR ≤135", "Восстановление"),
            date(2026, 6, 9): ("Пн: 2×800м @ темп быстрее 2.5К-pace + 20м разм + 15м зам",
                               "На 800м: 3:30-3:35 (~9:30 цель)", "Острая активация. Pre: + Гипоксен 500мг"),
            date(2026, 6, 10): ("Вт: 4-6×100м махом + 25 мин Z1", "Z1 ≤130", "Финальный нервно-мышечный звонок"),
            date(2026, 6, 11): ("Ср: отдых или 30 мин ходьба", "≤95", "Свежесть к старту"),
        }
        return plan_C.get(d, (plans_text, "—", "Свежесть, без интенсива"))

    return (plans_text, "—", "—")


def bads_for_day(d: date, phase_key: str) -> str:
    if phase_key == "A":
        return ("Утро (если тренировка): Цитрулин 6г + Карнитин (с мёдом). "
                "Без Женьшеня и без Гипоксена. "
                "Базовые: Цитофлавин 2×2 с едой, Omega+D3 завтрак, КардиоМагнил ужин, Цинк сон.")
    if phase_key == "B":
        if d == date(2026, 6, 2):
            return ("МПК-день! Полный pre-workout: Цитрулин + Карнитин + мёд + Женьшень + "
                    "**Гипоксен 500мг** + Стимол после")
        return ("Pre-workout полный (без Гипоксена кроме интенсива). "
                "Стимол после МПК. Остальное как обычно.")
    if phase_key == "C":
        if d == date(2026, 6, 9):
            return "Острая активация → Гипоксен 500мг за 30-45 мин + Стимол после"
        if d == date(2026, 6, 11):
            return "Вечер: Стимол 1 пакет (антиастеник перед стартом)"
        return "Только базовые (Цитрулин не обязателен)"
    if phase_key == "RACE":
        return ("Утро: вода + соль. За 60 мин: Цитрулин 6г + Гипоксен 500мг + Стимол. "
                "За 30 мин: Карнитин + банан. После: Glutamine + Цитофлавин")
    return "—"


def build_calendar(wb):
    ws = wb.active
    ws.title = "Календарь"

    headers = [
        "Дата", "День", "Фаза",
        "Из плана (xlsx)",
        "Рекомендация на день",
        "ЧСС-потолок",
        "БАДы",
        "Заметки",
        "Факт (заполнить)",
        "Feedback",
    ]
    widths = [10, 6, 18, 35, 40, 22, 50, 35, 30, 25]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 25

    # Расписание июня (xlsx не покрывает, описан здесь явно)
    june_plan = {
        date(2026, 6, 1): [("утро", "вело 1ч Z2 или восст. бег 45м")],
        date(2026, 6, 2): [("утро", "МПК Семёновский: 4-5×5'/8' R")],
        date(2026, 6, 3): [("утро", "Z1 30-45 мин или отдых (после МПК)")],
        date(2026, 6, 4): [("утро", "🔬 Лактат-ретест в лаборатории (ступенчатый)")],
        date(2026, 6, 5): [("утро", "Лёгкий Z1 30 мин (восст. после теста)")],
        date(2026, 6, 6): [("утро", "вело 1.5ч Z1-Z2 или бег 1ч Z2")],
        date(2026, 6, 7): [("утро", "A1 40 мин (объём -50% от пика)")],
        date(2026, 6, 8): [("утро", "A1 30 мин или отдых")],
        date(2026, 6, 9): [("утро", "Острая: 2×800м @ темп >2.5К-pace + 20м разм + 15м зам")],
        date(2026, 6, 10): [("утро", "4-6×100м махом + 25 мин Z1")],
        date(2026, 6, 11): [("утро", "Отдых или 30 мин ходьба")],
        date(2026, 6, 12): [("утро", "🏁 КОНТРОЛЬНАЯ 2.5К на максимум")],
    }

    start = date(2026, 5, 19)
    end = date(2026, 6, 12)
    row = 2
    d = start
    while d <= end:
        phase_label, phase_key = phase_for_date(d)
        sessions = for_date(d)
        # Из xlsx
        if sessions:
            plan_text = "\n".join(f"• {s.part}: {s.text}" for s in sessions)
        # Из моего гайда (для июньских дней xlsx пустой)
        elif d in june_plan:
            plan_text = "\n".join(f"• {part}: {text}" for part, text in june_plan[d])
            plan_text += "\n(вне xlsx — мой гайд)"
        else:
            plan_text = "—"
        rec, hr_cap, notes = recommendation_for_day(d, sessions, phase_key)
        bads = bads_for_day(d, phase_key)

        values = [
            d.strftime("%d.%m"),
            WEEKDAYS_RU[d.weekday()],
            phase_label,
            plan_text,
            rec,
            hr_cap,
            bads,
            notes,
            "",
            "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = ALIGN_WRAP
            cell.border = BORDER
            if phase_key == "A":
                cell.fill = FILL_PHASE_A
            elif phase_key == "B":
                cell.fill = FILL_PHASE_B
            elif phase_key == "C":
                cell.fill = FILL_PHASE_C
            elif phase_key == "RACE":
                cell.fill = FILL_RACE
                cell.font = FONT_BOLD

        # Высота строки — авто по содержимому
        max_lines = max(str(v).count("\n") + 1 for v in values if v)
        ws.row_dimensions[row].height = max(20, max_lines * 15)

        row += 1
        d += timedelta(days=1)

    # Заморозим первую строку
    ws.freeze_panes = "A2"


def build_decision_tree(wb):
    ws = wb.create_sheet("Утренний decision tree")
    headers = ["HRV (vs база 70)", "RHR (vs база 47-49)", "BB утро", "Решение"]
    widths = [22, 22, 18, 50]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 25

    rows = [
        ("≥65", "≤51", "≥70", "✅ По плану / xlsx (зелёный свет)"),
        ("55-65", "51-53", "50-70", "🟡 По плану, но ЧСС в нижней половине зоны"),
        ("45-55", "54-56", "30-50", "⚠️ Сократить объём на 30%, перейти из Z2 в Z1"),
        ("<45", "≥56", "<30", "❌ Отдых или 30 мин Z1 максимум"),
    ]
    fills = [
        PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        PatternFill(start_color="FFE5D0", end_color="FFE5D0", fill_type="solid"),
        PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
    ]
    for i, (r, f) in enumerate(zip(rows, fills), 2):
        for col, val in enumerate(r, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = ALIGN_WRAP
            cell.border = BORDER
            cell.fill = f
        ws.row_dimensions[i].height = 30

    # Уточнения
    ws.cell(row=7, column=1, value="Важные уточнения:").font = FONT_BOLD
    notes = [
        "• HRV провал на 20+ — СНАЧАЛА спроси: алкоголь? Поздний ужин? Стресс? Часы снимал? Симптомы простуды?",
        "• Температура ≥22°C → длительный бег ЗАМЕНИТЬ на вело (cardiovascular drift на беге значителен)",
        "• 3 дня подряд жёлтый/красный → пересборка недели",
        "• Стресс avg >25 без видимой причины — флаг недосна / алкоголя / болезни",
    ]
    for i, n in enumerate(notes, 8):
        ws.cell(row=i, column=1, value=n).alignment = ALIGN_WRAP
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
        ws.row_dimensions[i].height = 25


def build_bads_by_phase(wb):
    ws = wb.create_sheet("БАДы по фазам")
    headers = ["Фаза", "Pre-workout", "Базовые на день", "Особое"]
    widths = [18, 45, 45, 45]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 25

    rows = [
        ("A: Wind-in\n19-28.05",
         "Цитрулин 6г + Карнитин + мёд (без Женьшеня и Гипоксена)",
         "Цитофлавин 2×2 с едой, Omega+D3 с завтраком, КардиоМагнил после ужина, Цинк перед сном",
         "Glutamine 5г после каждой тренировки. Стимол только при провале HRV."),
        ("B: Раскачка\n29.05-06.06",
         "Полный: Цитрулин + Карнитин + мёд + Женьшень. Гипоксен — только в МПК-день (02.06).",
         "Те же базовые",
         "Стимол после МПК-сессии. Лактат-ретест 04-05.06: за день до — отдых, БАДов не дублировать."),
        ("C: Подводка\n07-11.06",
         "07-08: лёгкие. 09 (острая 2×800м): + Гипоксен 500мг. 10-11: только базовые.",
         "Те же базовые. Углеводная загрузка с 10.06 (+50-100 г/день)",
         "Стимол 1 пакет вечером 11.06 (антиастеник перед стартом)"),
        ("🏁 День старта\n12.06",
         "Утро вода+соль. За 60 мин: Цитрулин 6г + Гипоксен 500мг + Стимол. За 30 мин: Карнитин + банан.",
         "После старта: Glutamine + Цитофлавин + углеводы 60г",
         "Никаких новых препаратов! Только проверенный стек."),
    ]
    for i, r in enumerate(rows, 2):
        for col, val in enumerate(r, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = ALIGN_WRAP
            cell.border = BORDER
        ws.row_dimensions[i].height = 80


def build_tests_checklist(wb):
    ws = wb.create_sheet("Тесты + чек-листы")

    ws.cell(row=1, column=1, value="ТЕСТЫ ФОРМЫ").font = Font(bold=True, size=13)
    ws.row_dimensions[1].height = 22

    tests = [
        ["Тест", "Когда", "Что даёт", "Готово"],
        ["Семёновский ПН-тест (1 км ровно)", "каждый Пн утром", "2D-график: время + ЧСС 0/+1/+2 → тренд формы", ""],
        ["Лактат-ретест в лаборатории", "04 или 05.06", "Обновлённые LT1/LT2 → пересборка зон", ""],
        ["Тест 3×10 мин на ЧСС 130/140/150", "1 раз в начале июня + повтор в конце", "Скорость-при-HR динамика", ""],
        ["Контрольная 2.5К", "12.06", "Полевой LT2 + текущая беговая форма", ""],
    ]
    for i, r in enumerate(tests, 2):
        for col, val in enumerate(r, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = ALIGN_WRAP
            cell.border = BORDER
            if i == 2:  # header
                cell.fill = FILL_HEADER
                cell.font = FONT_HEADER
        ws.row_dimensions[i].height = 30 if i > 2 else 25

    for col, w in enumerate([35, 22, 50, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Кардио-чек-лист
    base_row = len(tests) + 5
    ws.cell(row=base_row, column=1, value="КАРДИО + ДИАГНОСТИКА").font = Font(bold=True, size=13)

    cardio = [
        ["Исследование", "Срок", "Статус", "Заметка"],
        ["ЭКГ", "до 25.05", "✅ Готово 18.05", "Норма для атлета"],
        ["Анализ крови (биохимия)", "до 25.05", "✅ Готово 18.05", "Норма. Альфа-амилаза 15 — мягко"],
        ["УЗИ брюшной + почки", "до 25.05", "✅ Готово 18.05", "Реактивная гепатомегалия, остальное норма"],
        ["ЭхоКГ", "до 01.06", "⏳ Записаться", "Толщина стенки ЛЖ + структура (важно)"],
        ["ОАК + коагулограмма", "до 15.06 (планово)", "⏳ Опционально", "Для оценки 'густой крови'"],
        ["Гастроэнтеролог (плановая)", "июнь", "⏳ Планово", "С УЗИ + биохимией"],
    ]
    for i, r in enumerate(cardio, base_row + 1):
        for col, val in enumerate(r, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = ALIGN_WRAP
            cell.border = BORDER
            if i == base_row + 1:
                cell.fill = FILL_HEADER
                cell.font = FONT_HEADER
        ws.row_dimensions[i].height = 28 if i > base_row + 1 else 25

    # Пред-стартовый чек-лист
    base_row2 = base_row + len(cardio) + 3
    ws.cell(row=base_row2, column=1, value="ПРЕД-СТАРТОВЫЙ ЧЕК-ЛИСТ (вечер 11.06)").font = Font(bold=True, size=13)

    checklist = [
        ["Пункт", "Статус"],
        ["Углеводная загрузка 10-11.06 (+50-100 г/день)", ""],
        ["Гидратация: ≥3 л воды + щепотка соли утром", ""],
        ["Снаряжение проверено (кроссовки, форма, чип)", ""],
        ["Сон 7-8 часов ночью с 11 на 12.06", ""],
        ["Завтрак за 2-3 ч (банан + овсянка + хлеб с мёдом)", ""],
        ["Pre-workout БАДы (Цитрулин + Гипоксен + Стимол)", ""],
        ["Разминка 15-20 мин: 5 мин трусцой + 5 мин махи + 5×80 м ускорения", ""],
        ["Прибыть к старту за 30-40 мин", ""],
    ]
    for i, r in enumerate(checklist, base_row2 + 1):
        for col, val in enumerate(r, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = ALIGN_WRAP
            cell.border = BORDER
            if i == base_row2 + 1:
                cell.fill = FILL_HEADER
                cell.font = FONT_HEADER
        ws.row_dimensions[i].height = 25


def main():
    wb = Workbook()
    build_calendar(wb)
    build_decision_tree(wb)
    build_bads_by_phase(wb)
    build_tests_checklist(wb)

    out_path = Path(__file__).resolve().parent.parent / "plans" / "corrections_to_12-06-2026.xlsx"
    wb.save(str(out_path))
    print(f"Сохранено: {out_path}")
    print(f"Размер: {out_path.stat().st_size} байт")


if __name__ == "__main__":
    main()
