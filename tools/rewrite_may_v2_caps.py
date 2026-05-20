"""Полная переписка W2-W5 в Excel v2 под новые лимиты.

Лимиты:
- Утро Пн-Пт: ≤ 1:15
- Сб: до 3:00
- Вс: до 2:30
- Май = втягивающий блок, ~52 ч итого
"""
from __future__ import annotations

import os
import sys

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

dst = "plans/2026_05_may_schedule_v2.xlsx"
wb = openpyxl.load_workbook(dst, data_only=False)

# Колонки в Календаре: Пн=2, Вт=3, Ср=4, Чт=5, Пт=6, Сб=7, Вс=8
# Структура: row N — заголовки дат, N+1 — утро, N+2 — вечер

# === Содержимое W2-W5 ===
# format: (week_label_match, hours_total, [(col, morning, evening, dur_h)])
WEEK_DATA = {
    "W2": {
        "hours": 11.5,
        "morning": {
            2: ("☀ бег 45м Z1 🛣️ шоссе (восст) + 🤸 кор 16м", "бег"),
            3: ("☀ вело 1ч Z2", "вело"),
            4: ("☀ вело 1ч 15м Z2", "вело"),
            5: ("🦘 прыжковая беговая 45м 🌲 пересеч/грунт (нейромыш. активация)", "плиометрика"),
            6: ("☀ вело 1ч Z2 + 4×30 сек В ПОДЪЁМ", "вело"),
            7: ("☀ длит. вело 2ч 30м Z1-Z2 🌲 рельеф", "вело-длит"),
            8: ("☀ бег 2ч Z2 🌲 пересеч", "бег-длит"),
        },
        "evening": {
            3: ("🌙 🏋 СТ ОМВ-НИЗ ЗАЛ 1ч (3 круга, втягивающий)", "ст-омв"),
            6: ("🌙 🤸 СТ ОМВ-ВЕРХ УЛИЦА 1ч (3 круга, втягивающий)", "ст-омв"),
        },
    },
    "W3": {
        "hours": 12.5,
        "morning": {
            2: ("☀ бег 1ч Z1 🛣️ шоссе", "бег"),
            3: ("☀ вело 1ч Z2", "вело"),
            4: ("☀ вело 1ч 15м Z2 + 3-4×20 сек в короткий подъём", "вело"),
            5: ("☀ бег 30м Z1 🛣️ + 🤸 кор 16м + растяжка 15м", "бег"),
            6: ("☀ вело 1ч Z2", "вело"),
            7: ("☀ длит. вело 2ч 45м Z1-Z2 🌲 рельеф", "вело-длит"),
            8: ("☀ бег 2ч Z2 🌲 пересеч", "бег-длит"),
        },
        "evening": {
            3: ("🌙 🏋 СТ ОМВ-НИЗ ЗАЛ 1ч 15м (4 круга развивающий)", "ст-омв"),
            6: ("🌙 🤸 СТ ОМВ-ВЕРХ УЛИЦА 1ч 15м (4 круга развивающий)", "ст-омв"),
        },
    },
    "W4": {
        "hours": 13.25,
        "morning": {
            2: ("☀ бег 1ч Z1 🛣️ шоссе (восст)", "бег"),
            3: ("☀ вело 1ч Z2", "вело"),
            4: ("☀ вело 1ч 15м Z2 + 2×10м Z2-верх АэП~155-160", "вело"),
            5: ("☀ бег 45м Z1 🛣️ + 🤸 кор 16м", "бег"),
            6: ("☀ вело 1ч Z2 + 5×30 сек В ПОДЪЁМ", "вело"),
            7: ("☀ длит. вело 3ч Z1-Z2 🌲 рельеф (макс мая)", "вело-длит"),
            8: ("☀ бег 2ч 30м Z2 🌲 пересеч (макс мая)", "бег-длит"),
        },
        "evening": {
            3: ("🌙 🏋 СТ ОМВ-НИЗ ЗАЛ 1ч 15м (5 кругов ПИК)", "ст-омв-пик"),
            6: ("🌙 🤸 СТ ОМВ-ВЕРХ УЛИЦА 1ч 15м (5 кругов ПИК)", "ст-омв-пик"),
        },
    },
    "W5": {
        "hours": 8.5,
        "morning": {
            2: ("☀ отдых ИЛИ прогулка 30м", "отдых"),
            3: ("☀ вело 1ч Z1 (восст)", "вело"),
            4: ("☀ бег 1ч Z2 🛣️ шоссе (плавный возврат)", "бег"),
            5: ("☀ вело 1ч Z2 + 🤸 кор 16м", "вело"),
            6: ("☀ бег 1ч Z2 🛣️ шоссе + 4×30 сек разгон", "бег"),
            7: ("☀ вело 2ч Z2", "вело-длит"),
            8: ("☀ 🏁 контр. 5К 🛣️ шоссе: 30м разм + 5К + 15м зам", "контроль"),
        },
        "evening": {
            3: ("🌙 🤸 СТ ОМВ ТОНИЗИР. УЛИЦА 45м (1-2 круга легко)", "ст-тон"),
        },
    },
}

# Часы по дням (для Тренировки списком)
HOURS_PER_DAY = {
    # W2
    "2026-05-04": {"утро": 1.0},  # бег 45м + кор 16м
    "2026-05-05": {"утро": 1.0, "вечер": 1.0},
    "2026-05-06": {"утро": 1.25},
    "2026-05-07": {"утро": 0.75},
    "2026-05-08": {"утро": 1.0, "вечер": 1.0},
    "2026-05-09": {"утро": 2.5},
    "2026-05-10": {"утро": 2.0},
    # W3
    "2026-05-11": {"утро": 1.0},
    "2026-05-12": {"утро": 1.0, "вечер": 1.25},
    "2026-05-13": {"утро": 1.25},
    "2026-05-14": {"утро": 1.0},
    "2026-05-15": {"утро": 1.0, "вечер": 1.25},
    "2026-05-16": {"утро": 2.75},
    "2026-05-17": {"утро": 2.0},
    # W4
    "2026-05-18": {"утро": 1.0},
    "2026-05-19": {"утро": 1.0, "вечер": 1.25},
    "2026-05-20": {"утро": 1.25},
    "2026-05-21": {"утро": 1.0},
    "2026-05-22": {"утро": 1.0, "вечер": 1.25},
    "2026-05-23": {"утро": 3.0},
    "2026-05-24": {"утро": 2.5},
    # W5
    "2026-05-25": {"утро": 0.0},
    "2026-05-26": {"утро": 1.0, "вечер": 0.75},
    "2026-05-27": {"утро": 1.0},
    "2026-05-28": {"утро": 1.25},
    "2026-05-29": {"утро": 1.0},
    "2026-05-30": {"утро": 2.0},
    "2026-05-31": {"утро": 1.5},
}

# === Календарь ===
ws = wb["Календарь"]
print("=== Календарь ===")
for r in range(1, ws.max_row + 1):
    label = ws.cell(row=r, column=1).value
    if not label:
        continue
    label_str = str(label)
    for week_key, data in WEEK_DATA.items():
        if week_key in label_str:
            morning_row = r + 1
            evening_row = r + 2
            # Сначала очистим вечер для дней которые мы НЕ переписываем (на всякий случай)
            # Утро всегда есть в нашем словаре для всех 7 дней
            for col, (text, _typ) in data["morning"].items():
                ws.cell(row=morning_row, column=col).value = text
            # Вечер — заменяем что в словаре, остальные оставляем "—"
            for col in range(2, 9):
                if col in data["evening"]:
                    text, _typ = data["evening"][col]
                    ws.cell(row=evening_row, column=col).value = text
                else:
                    # Только если в текущей ячейке нет полезной инфы — ставим тире
                    cur = ws.cell(row=evening_row, column=col).value
                    if not cur or cur == "—" or "—" in str(cur):
                        ws.cell(row=evening_row, column=col).value = "—"
                    else:
                        # Очищаем устаревшее (старые СТ, кор и т.д.)
                        ws.cell(row=evening_row, column=col).value = "—"
            ws.cell(row=r, column=9).value = data["hours"]
            print(f"  {week_key}: обновлено, часы={data['hours']}")
            break

# Итого май
total = sum(d["hours"] for d in WEEK_DATA.values()) + 6.7  # +W1 факт
total = round(total, 1)
for r in range(1, ws.max_row + 1):
    if "ИТОГО" in str(ws.cell(row=r, column=1).value or ""):
        ws.cell(row=r, column=9).value = total
        print(f"  ИТОГО: {total}")
        break

# === Тренировки списком ===
ws2 = wb["Тренировки списком"]
print("\n=== Тренировки списком — пересоздание ===")
# Находим headers и удаляем все строки данных
header_row = 1  # должен быть в строке 1
# Удалим все строки ниже header — будем переписывать с нуля
max_r = ws2.max_row
if max_r > 1:
    ws2.delete_rows(2, max_r - 1)

# Для каждой даты W1-W5 формируем строки (W1 = факт, оставляем как было)
W1_FACT = [
    ("2026-05-01", "Пт", 1, "утро", "✅ бег 19.06км / 1:51 / HR 134 (+228м)", "бег-длит", 1.85),
    ("2026-05-01", "Пт", 1, "вечер", "✅ вело 1.5км / 4м (перемещение)", "вело", 0.07),
    ("2026-05-02", "Сб", 1, "утро", "✅ вело 60.57км / 2:21 / HR 130 (+207м)", "вело-длит", 2.35),
    ("2026-05-02", "Сб", 1, "вечер", "✅ вело 3км / 9м (перемещение)", "вело", 0.15),
    ("2026-05-03", "Вс", 1, "утро", "✅ вело 60.43км / 2:19 / HR 128 (+216м)", "вело-длит", 2.32),
]

DAY_SHORT = {0: "Пн", 1: "Вт", 2: "Ср", 3: "Чт", 4: "Пт", 5: "Сб", 6: "Вс"}
COL_TO_DAY = {2: 0, 3: 1, 4: 2, 5: 3, 6: 4, 7: 5, 8: 6}  # день недели по колонке

import datetime
def date_for(week_num: int, weekday: int) -> str:
    """W2 начинается с 04.05 (Пн), W3 = 11.05, W4 = 18.05, W5 = 25.05."""
    base = {2: 4, 3: 11, 4: 18, 5: 25}
    day = base[week_num] + weekday
    if day <= 31:
        return f"2026-05-{day:02d}"
    return f"2026-06-{day-31:02d}"

rows_to_add = list(W1_FACT)
for week_key, data in WEEK_DATA.items():
    week_num = int(week_key[1:])
    for col, (text, typ) in data["morning"].items():
        weekday = COL_TO_DAY[col]
        date_str = date_for(week_num, weekday)
        # Текст без префикса ☀ для списка
        clean = text.replace("☀ ", "").replace("🌙 ", "")
        hours = HOURS_PER_DAY.get(date_str, {}).get("утро", 0)
        rows_to_add.append((date_str, DAY_SHORT[weekday], week_num, "утро", clean, typ, hours))
    for col, (text, typ) in data["evening"].items():
        weekday = COL_TO_DAY[col]
        date_str = date_for(week_num, weekday)
        clean = text.replace("☀ ", "").replace("🌙 ", "")
        hours = HOURS_PER_DAY.get(date_str, {}).get("вечер", 0)
        rows_to_add.append((date_str, DAY_SHORT[weekday], week_num, "вечер", clean, typ, hours))

# Сортируем по дате+утро/вечер
rows_to_add.sort(key=lambda x: (x[0], 0 if x[3] == "утро" else 1))

# Записываем
for i, row in enumerate(rows_to_add, start=2):
    for j, val in enumerate(row, start=1):
        ws2.cell(row=i, column=j).value = val
print(f"  Записано {len(rows_to_add)} строк")

# === Сводка ===
ws3 = wb["Сводка"]
print("\n=== Сводка ===")
# По неделям
week_hours = {"W1": 6.7}
week_hours.update({k: v["hours"] for k, v in WEEK_DATA.items()})
for r in range(5, 11):
    label = ws3.cell(row=r, column=1).value
    if not label:
        continue
    label_str = str(label)
    for k, h in week_hours.items():
        if k in label_str:
            ws3.cell(row=r, column=2).value = h
            break
    if "ИТОГО" in label_str:
        ws3.cell(row=r, column=2).value = total

# Очищаем строки типов 14-35 и пересчитываем
for r in range(15, 36):
    for c in range(1, 4):
        ws3.cell(row=r, column=c).value = None
ws3.cell(row=14, column=1).value = "Тип"
ws3.cell(row=14, column=2).value = "Часы"
ws3.cell(row=14, column=3).value = "% от месяца"

type_totals: dict[str, float] = {}
for row in rows_to_add:
    typ = row[5]
    h = float(row[6])
    type_totals[typ] = type_totals.get(typ, 0) + h

r = 15
for t, h in sorted(type_totals.items(), key=lambda x: -x[1]):
    ws3.cell(row=r, column=1).value = t
    ws3.cell(row=r, column=2).value = round(h, 2)
    ws3.cell(row=r, column=3).value = f"{h / total * 100:.1f}%"
    r += 1
print(f"  Типы записаны, итого {total}")

wb.save(dst)
print(f"\nSaved: {dst}")
print(f"Size: {os.path.getsize(dst):,} bytes")
print(f"Total month hours: {total}")
