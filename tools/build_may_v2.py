"""Создать обновлённую версию плана мая в Excel с правками от 2026-05-03.

Изменения:
- W2 Пн 04.05: + кор 16м к утреннему бегу
- W2 Чт 07.05: бег + растяжка → прыжковая беговая 40-45м
- W3 Чт 14.05: вечером + кор 16м + растяжка 15м (вместо одной растяжки 30м)
- W4 Чт 21.05: бег + растяжка → бег + кор 16м
- W5 Чт 28.05: вечером + кор 16м
"""
from __future__ import annotations

import os
import shutil
import sys

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

src = "plans/2026_05_may_schedule.xlsx"
dst = "plans/2026_05_may_schedule_v2.xlsx"
shutil.copy(src, dst)
print(f"Copied: {src} -> {dst}")

wb = openpyxl.load_workbook(dst, data_only=False)

# === Календарь ===
ws = wb["Календарь"]


def find_week_rows(ws):
    rows = {}
    for r in range(1, ws.max_row + 1):
        for c in range(2, 9):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str):
                if v.startswith("Пн ") or v.startswith("Пт 01.05"):
                    week_label = ws.cell(row=r, column=1).value
                    rows[str(week_label)] = (r + 1, r + 2, r)
                    break
    return rows


week_rows = find_week_rows(ws)
print("Найдены недели:", list(week_rows.keys()))

w2_key = next(k for k in week_rows if "W2" in k)
w3_key = next(k for k in week_rows if "W3" in k)
w4_key = next(k for k in week_rows if "W4" in k)
w5_key = next(k for k in week_rows if "W5" in k)

w2_m, w2_e, w2_h = week_rows[w2_key]
w3_m, w3_e, w3_h = week_rows[w3_key]
w4_m, w4_e, w4_h = week_rows[w4_key]
w5_m, w5_e, w5_h = week_rows[w5_key]


def set_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    old = cell.value
    cell.value = value
    return old


# Колонки: Пн=2, Вт=3, Ср=4, Чт=5, Пт=6, Сб=7, Вс=8
old = set_cell(ws, w2_m, 2, "☀ бег 1ч Z1 (восст) + \U0001F938 кор-стабилизация 16м")
print(f"W2 Пн утро: {old!r} -> updated")

old = set_cell(ws, w2_m, 5, "\U0001F998 прыжковая беговая 40-45м (нейромыш. активация)")
print(f"W2 Чт утро: {old!r} -> updated")

old = set_cell(ws, w3_e, 5, "\U0001F319 \U0001F938 кор-стабилизация 16м + растяжка 15м")
print(f"W3 Чт вечер: {old!r} -> updated")

old = set_cell(ws, w4_m, 5, "☀ бег 1ч Z1 + \U0001F938 кор-стабилизация 16м")
print(f"W4 Чт утро: {old!r} -> updated")

old = set_cell(ws, w5_e, 5, "\U0001F319 \U0001F938 кор-стабилизация 16м")
print(f"W5 Чт вечер: {old!r} -> updated")

# Часы по неделям в колонке 9
new_hours = {"W2": 14.1, "W3": 15.2, "W4": 16.0, "W5": 11.1}
total_new = 5.6 + sum(new_hours.values())  # 5.6 + 56.4 = 62.0
for r in range(1, ws.max_row + 1):
    label = ws.cell(row=r, column=1).value
    if not label:
        continue
    label = str(label)
    for w, h in new_hours.items():
        if w in label:
            ws.cell(row=r, column=9).value = h
            break
    if "ИТОГО" in label:
        ws.cell(row=r, column=9).value = round(total_new, 1)

# === Тренировки списком ===
ws2 = wb["Тренировки списком"]


def update_row(ws, date_str, day_part, new_text=None, new_type=None, new_hours=None):
    for r in range(2, ws.max_row + 1):
        d = ws.cell(row=r, column=1).value
        p = ws.cell(row=r, column=4).value
        if str(d).startswith(date_str) and p == day_part:
            if new_text is not None:
                ws.cell(row=r, column=5).value = new_text
            if new_type is not None:
                ws.cell(row=r, column=6).value = new_type
            if new_hours is not None:
                ws.cell(row=r, column=7).value = new_hours
            print(f"  Updated row: {date_str} {day_part}")
            return r
    return None


update_row(
    ws2,
    "2026-05-04",
    "утро",
    new_text="бег 1ч Z1 (восст) + \U0001F938 кор-стабилизация 16м",
    new_hours=1.27,
)

update_row(
    ws2,
    "2026-05-07",
    "утро",
    new_text="\U0001F998 прыжковая беговая 40-45м (нейромыш. активация)",
    new_type="плиометрика",
    new_hours=0.75,
)

update_row(
    ws2,
    "2026-05-21",
    "утро",
    new_text="бег 1ч Z1 + \U0001F938 кор-стабилизация 16м",
    new_hours=1.27,
)


def insert_evening_after(ws, date_str, week_num, text, type_, hours):
    for r in range(2, ws.max_row + 1):
        d = ws.cell(row=r, column=1).value
        p = ws.cell(row=r, column=4).value
        if str(d).startswith(date_str) and p == "утро":
            ws.insert_rows(r + 1)
            ws.cell(row=r + 1, column=1).value = ws.cell(row=r, column=1).value
            ws.cell(row=r + 1, column=2).value = ws.cell(row=r, column=2).value
            ws.cell(row=r + 1, column=3).value = week_num
            ws.cell(row=r + 1, column=4).value = "вечер"
            ws.cell(row=r + 1, column=5).value = text
            ws.cell(row=r + 1, column=6).value = type_
            ws.cell(row=r + 1, column=7).value = hours
            print(f"  Inserted evening row: {date_str}")
            return


insert_evening_after(
    ws2,
    "2026-05-14",
    3,
    "\U0001F938 кор-стабилизация 16м + растяжка 15м",
    "кор",
    0.5,
)

insert_evening_after(
    ws2,
    "2026-05-28",
    5,
    "\U0001F938 кор-стабилизация 16м",
    "кор",
    0.27,
)

# === Сводка ===
ws3 = wb["Сводка"]
for r in range(5, 11):
    label = ws3.cell(row=r, column=1).value
    if label is None:
        continue
    label = str(label)
    for w, h in new_hours.items():
        if w in label:
            ws3.cell(row=r, column=2).value = h
            break
    if "ИТОГО" in label:
        ws3.cell(row=r, column=2).value = round(total_new, 1)

# Типы тренировок: бег уменьшен на 1ч (Чт 07.05 заменён)
type_updates = {"бег": 9.6}
new_types = [("плиометрика", 0.75), ("кор", 1.04)]

for r in range(15, 24):
    label = ws3.cell(row=r, column=1).value
    if label in type_updates:
        new_h = type_updates[label]
        ws3.cell(row=r, column=2).value = new_h
        ws3.cell(row=r, column=3).value = f"{new_h / total_new * 100:.1f}%"

last_used_row = 23
for name, h in new_types:
    last_used_row += 1
    ws3.cell(row=last_used_row, column=1).value = name
    ws3.cell(row=last_used_row, column=2).value = h
    ws3.cell(row=last_used_row, column=3).value = f"{h / total_new * 100:.1f}%"

wb.save(dst)
print(f"\nSaved: {dst}")
print(f"Size: {os.path.getsize(dst):,} bytes")
