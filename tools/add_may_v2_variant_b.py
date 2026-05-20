"""Вариант B: добавить 2-ю беговую/велo Ср+Чт вечер в W3 и W4 в Excel v2."""
from __future__ import annotations

import os
import sys

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

dst = "plans/2026_05_may_schedule_v2.xlsx"
wb = openpyxl.load_workbook(dst, data_only=False)

# === Календарь ===
ws = wb["Календарь"]

# Колонки: Ср=4, Чт=5
# Структура: row N — заголовки дат, N+1 — утро, N+2 — вечер
# Обновим часы в заголовках W3 и W4 + добавим вечерние записи

ADDITIONS = {
    "W3": {
        "evening": {
            4: "🌙 ☀ бег 1ч Z1 🛣️ шоссе (легко, 2-я беговая)",  # Ср
            5: "🌙 вело 1ч Z1 (актив. восст.)",  # Чт
        },
        "new_hours": 14.5,
    },
    "W4": {
        "evening": {
            4: "🌙 ☀ бег 1ч Z1 🛣️ шоссе (легко, 2-я беговая)",
            5: "🌙 вело 1ч Z1 (актив. восст.)",
        },
        "new_hours": 15.25,
    },
}

print("=== Календарь ===")
for r in range(1, ws.max_row + 1):
    label = ws.cell(row=r, column=1).value
    if not label:
        continue
    label_str = str(label)
    for week_key, data in ADDITIONS.items():
        if week_key in label_str:
            evening_row = r + 2
            for col, text in data["evening"].items():
                # Заменяем "—" на новую сессию
                ws.cell(row=evening_row, column=col).value = text
                day_name = {4: "Ср", 5: "Чт"}[col]
                print(f"  {week_key} {day_name} вечер: {text}")
            ws.cell(row=r, column=9).value = data["new_hours"]
            print(f"  {week_key} часы: {data['new_hours']}")
            break

# Итого май: было 52.5, новое 56.5
new_total = 56.5
for r in range(1, ws.max_row + 1):
    if "ИТОГО" in str(ws.cell(row=r, column=1).value or ""):
        ws.cell(row=r, column=9).value = new_total
        print(f"  ИТОГО: {new_total}")
        break

# === Тренировки списком — добавляем 4 новые строки ===
ws2 = wb["Тренировки списком"]
print("\n=== Тренировки списком ===")
NEW_ROWS = [
    ("2026-05-13", "Ср", 3, "вечер", "бег 1ч Z1 🛣️ шоссе (легко, 2-я беговая)", "бег", 1.0),
    ("2026-05-14", "Чт", 3, "вечер", "вело 1ч Z1 (актив. восст.)", "вело", 1.0),
    ("2026-05-20", "Ср", 4, "вечер", "бег 1ч Z1 🛣️ шоссе (легко, 2-я беговая)", "бег", 1.0),
    ("2026-05-21", "Чт", 4, "вечер", "вело 1ч Z1 (актив. восст.)", "вело", 1.0),
]

# Найдём пустую строку в конце листа и просто добавим
last_row = ws2.max_row
for new_row in NEW_ROWS:
    last_row += 1
    for j, val in enumerate(new_row, start=1):
        ws2.cell(row=last_row, column=j).value = val
    print(f"  +строка: {new_row[0]} {new_row[3]} {new_row[4]}")

# Пересортируем — соберём все строки данных, отсортируем, перепишем
print("  Пересортирую по дате+утро/вечер...")
all_rows = []
for r in range(2, ws2.max_row + 1):
    row_data = tuple(ws2.cell(row=r, column=c).value for c in range(1, 8))
    if row_data[0]:
        all_rows.append(row_data)

# Очищаем
for r in range(2, ws2.max_row + 1):
    for c in range(1, 8):
        ws2.cell(row=r, column=c).value = None

# Сортируем
all_rows.sort(key=lambda x: (str(x[0]), 0 if x[3] == "утро" else 1))

# Записываем
for i, row in enumerate(all_rows, start=2):
    for j, val in enumerate(row, start=1):
        ws2.cell(row=i, column=j).value = val
print(f"  Всего строк: {len(all_rows)}")

# === Сводка ===
ws3 = wb["Сводка"]
print("\n=== Сводка ===")
# Обновим W3, W4, ИТОГО
for r in range(5, 11):
    label = ws3.cell(row=r, column=1).value
    if not label:
        continue
    label_str = str(label)
    if "W3" in label_str:
        ws3.cell(row=r, column=2).value = 14.5
    elif "W4" in label_str:
        ws3.cell(row=r, column=2).value = 15.25
    elif "ИТОГО" in label_str:
        ws3.cell(row=r, column=2).value = new_total

# Пересчёт по типам
type_totals: dict[str, float] = {}
for row in all_rows:
    typ = row[5]
    h = float(row[6]) if row[6] is not None else 0
    type_totals[typ] = type_totals.get(typ, 0) + h

# Очищаем строки типов 15-35 и пишем заново
for r in range(15, 36):
    for c in range(1, 4):
        ws3.cell(row=r, column=c).value = None

r = 15
for t, h in sorted(type_totals.items(), key=lambda x: -x[1]):
    if t and h > 0:
        ws3.cell(row=r, column=1).value = t
        ws3.cell(row=r, column=2).value = round(h, 2)
        ws3.cell(row=r, column=3).value = f"{h / new_total * 100:.1f}%"
        r += 1
print(f"  Типы пересчитаны (всего {len(type_totals)} типов, итого {new_total} ч)")

wb.save(dst)
print(f"\nSaved: {dst}")
print(f"Size: {os.path.getsize(dst):,} bytes")
print(f"Total month: {new_total} ч")
