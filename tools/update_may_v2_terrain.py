"""Добавить тэги покрытия (🛣️ шоссе / 🌲 пересеченка) к беговым тренировкам в Excel v2."""
from __future__ import annotations

import os
import sys

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

dst = "plans/2026_05_may_schedule_v2.xlsx"
wb = openpyxl.load_workbook(dst, data_only=False)

ROAD = "🛣️"
TRAIL = "🌲"

# Маппинг (date_str, day_part) -> новый текст
# Брал из текущего содержимого Excel, добавил тэги
calendar_updates = {
    # ====== W2 ======
    # Пн 04.05 утро: бег 1ч Z1 (восст активн.) + кор -- но кор уже пристроен в v2
    ("2026-05-04", "утро"): f"☀ бег 1ч Z1 {ROAD} шоссе (восст) + 🤸 кор-стабилизация 16м",
    # Чт 07.05: прыжковая
    ("2026-05-07", "утро"): f"🦘 прыжковая беговая 40-45м {TRAIL} пересеченка/грунт (нейромыш. активация)",
    # Вс 10.05: бег 2ч 15м Z2 (рельеф) — длинная
    ("2026-05-10", "утро"): f"☀ бег 2ч 15м Z2 {TRAIL} пересеченка — длинная",
    # ====== W3 ======
    # Пн 11.05: бег 1ч 15м Z1
    ("2026-05-11", "утро"): f"☀ бег 1ч 15м Z1 {ROAD} шоссе",
    # Чт 14.05 утро: бег 1ч Z1 + растяжка
    ("2026-05-14", "утро"): f"☀ бег 1ч Z1 {ROAD} шоссе + растяжка",
    # Вс 17.05: бег 2ч 20м Z2 (рельеф) — длинная
    ("2026-05-17", "утро"): f"☀ бег 2ч 20м Z2 {TRAIL} пересеченка — длинная",
    # ====== W4 ======
    # Пн 18.05: бег 1ч 20м Z1 (восст)
    ("2026-05-18", "утро"): f"☀ бег 1ч 20м Z1 {ROAD} шоссе (восст)",
    # Чт 21.05 утро: бег 1ч Z1 + кор
    ("2026-05-21", "утро"): f"☀ бег 1ч Z1 {ROAD} шоссе + 🤸 кор-стабилизация 16м",
    # Вс 24.05: бег 2ч 30м Z2 (рельеф) — длинная мая
    ("2026-05-24", "утро"): f"☀ бег 2ч 30м Z2 {TRAIL} пересеченка — длинная мая",
    # ====== W5 ======
    # Ср 27.05: бег 1ч 30м Z2 (плавный возврат)
    ("2026-05-27", "утро"): f"☀ бег 1ч 30м Z2 {ROAD} шоссе (плавный возврат)",
    # Пт 29.05: бег 1ч 15м Z2 + 4×30 сек разгон НА РАВНИНЕ
    ("2026-05-29", "утро"): f"☀ бег 1ч 15м Z2 {ROAD} шоссе + 4×30 сек разгон",
    # Вс 31.05: контр 5К
    ("2026-05-31", "утро"): f"☀ 🏁 контр. 5К {ROAD} шоссе: 30м разм + 5К + 15м зам",
}

# === Тренировки списком ===
ws2 = wb["Тренировки списком"]
print("=== Тренировки списком ===")
for r in range(2, ws2.max_row + 1):
    d = ws2.cell(row=r, column=1).value
    p = ws2.cell(row=r, column=4).value
    if not d or not p:
        continue
    key = (str(d), p)
    if key in calendar_updates:
        old = ws2.cell(row=r, column=5).value
        # Для tренировок-списком обрезаем символы солнца/луны вначале
        new_text = calendar_updates[key]
        # Уберём префикс "☀ " или "🌙 " если есть
        if new_text.startswith("☀ "):
            new_text = new_text[2:]
        elif new_text.startswith("🌙 "):
            new_text = new_text[2:]
        ws2.cell(row=r, column=5).value = new_text
        print(f"  {d} {p}: {old!r} -> {new_text!r}")

# === Календарь ===
ws = wb["Календарь"]


def find_date_cell(ws, date_str):
    """Найти ячейку с заголовком даты в Календаре. Возвращает (row, col)."""
    # Дата в Календаре в формате "Пн 04.05" — для каждого дня недели
    target_dm = date_str[5:].replace("-", ".")  # 2026-05-04 -> 05.04 (нужно 04.05)
    target_dm = ".".join(reversed(target_dm.split(".")))  # 05.04 -> 04.05
    for r in range(1, ws.max_row + 1):
        for c in range(2, 9):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str) and target_dm in v:
                return r, c
    return None, None


print("\n=== Календарь ===")
for (date_str, day_part), new_text in calendar_updates.items():
    r, c = find_date_cell(ws, date_str)
    if r is None:
        print(f"  WARNING: not found in Календарь: {date_str}")
        continue
    target_row = r + 1 if day_part == "утро" else r + 2
    old = ws.cell(row=target_row, column=c).value
    ws.cell(row=target_row, column=c).value = new_text
    print(f"  {date_str} {day_part} (r={target_row},c={c}): {old!r} -> {new_text!r}")

wb.save(dst)
print(f"\nSaved: {dst}")
print(f"Size: {os.path.getsize(dst):,} bytes")
