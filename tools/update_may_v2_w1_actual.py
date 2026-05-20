"""Обновить v2 Excel плана мая: заменить W1 хвост (1-3.05) с плана на факт по Garmin."""
from __future__ import annotations

import os
import sys

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

dst = "plans/2026_05_may_schedule_v2.xlsx"
wb = openpyxl.load_workbook(dst, data_only=False)

# === Календарь: W1 хвост ===
ws = wb["Календарь"]


def find_week_rows(ws):
    rows = {}
    for r in range(1, ws.max_row + 1):
        for c in range(2, 9):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str):
                if v.startswith("Пн ") or v.startswith("Пт 01.05"):
                    week_label = ws.cell(row=r, column=1).value
                    rows[str(week_label)] = (r + 1, r + 2, r)
                    break
    return rows


week_rows = find_week_rows(ws)
w1_key = next(k for k in week_rows if "W1" in k)
w1_morning, w1_evening, w1_header = week_rows[w1_key]

# Обновляем заголовок W1 — добавляем «✅ факт»
old_label = ws.cell(row=w1_header, column=1).value
new_label = "W1 хвост\n✅ ФАКТ 6.7ч (план 4ч, +65-70%)"
ws.cell(row=w1_header, column=1).value = new_label
print(f"W1 заголовок: {old_label!r} -> {new_label!r}")

# Колонки: Пт=6, Сб=7, Вс=8
# Утро (w1_morning):
ws.cell(row=w1_morning, column=6).value = (
    "✅ \U0001F3C3 бег 19.06км / 1:51 / HR 134 / +228м"
)
ws.cell(row=w1_morning, column=7).value = (
    "✅ \U0001F6B4 вело 60.57км / 2:21 / HR 130 / +207м"
)
ws.cell(row=w1_morning, column=8).value = (
    "✅ \U0001F6B4 вело 60.43км / 2:19 / HR 128 / +216м"
)

# Вечер (w1_evening): добавляем короткие вечерние/перемещения
ws.cell(row=w1_evening, column=6).value = "🚴 вело 4м перемещение"
ws.cell(row=w1_evening, column=7).value = "🚴 вело 9м перемещение"
ws.cell(row=w1_evening, column=8).value = "—"

# Часы W1: 5.6 -> 6.7
ws.cell(row=w1_header, column=9).value = 6.7

# ИТОГО май: 62 -> 63.1
for r in range(1, ws.max_row + 1):
    label = ws.cell(row=r, column=1).value
    if label and "ИТОГО" in str(label):
        ws.cell(row=r, column=9).value = 63.1
        break

# === Тренировки списком ===
ws2 = wb["Тренировки списком"]

# Обновляем существующие строки 01-03.05 + добавляем недостающие
# План был:
#   2026-05-01 Пт утро: вело 1ч + 4×30 (1h)
#   2026-05-02 Сб утро: длит вело 3:20 (3.33h)
#   2026-05-03 Вс утро: бег 1:15 рельеф (1.25h)

# Заменяем содержимое:
def update_or_get(ws, date_str, day_part):
    for r in range(2, ws.max_row + 1):
        d = ws.cell(row=r, column=1).value
        p = ws.cell(row=r, column=4).value
        if str(d).startswith(date_str) and p == day_part:
            return r
    return None


# 01.05 утро: бег
r = update_or_get(ws2, "2026-05-01", "утро")
ws2.cell(row=r, column=5).value = "✅ бег 19.06км / 1:51 / HR 134 (+228м)"
ws2.cell(row=r, column=6).value = "бег-длит"
ws2.cell(row=r, column=7).value = 1.85

# 01.05 + добавим вело перемещение (вторая запись в день — пометим как «вечер»
# хотя по факту ~12:34, но в схеме списка только утро/вечер)
ws2.insert_rows(r + 1)
ws2.cell(row=r + 1, column=1).value = ws2.cell(row=r, column=1).value
ws2.cell(row=r + 1, column=2).value = ws2.cell(row=r, column=2).value
ws2.cell(row=r + 1, column=3).value = 1
ws2.cell(row=r + 1, column=4).value = "вечер"
ws2.cell(row=r + 1, column=5).value = "✅ вело 1.5км / 4м (перемещение)"
ws2.cell(row=r + 1, column=6).value = "вело"
ws2.cell(row=r + 1, column=7).value = 0.07

# 02.05 утро: вело
r = update_or_get(ws2, "2026-05-02", "утро")
ws2.cell(row=r, column=5).value = "✅ вело 60.57км / 2:21 / HR 130 (+207м)"
ws2.cell(row=r, column=7).value = 2.35

# 02.05 + вечер вело 9м (новая строка)
ws2.insert_rows(r + 1)
ws2.cell(row=r + 1, column=1).value = ws2.cell(row=r, column=1).value
ws2.cell(row=r + 1, column=2).value = ws2.cell(row=r, column=2).value
ws2.cell(row=r + 1, column=3).value = 1
ws2.cell(row=r + 1, column=4).value = "вечер"
ws2.cell(row=r + 1, column=5).value = "✅ вело 3км / 9м (перемещение)"
ws2.cell(row=r + 1, column=6).value = "вело"
ws2.cell(row=r + 1, column=7).value = 0.15

# 03.05 утро: вело (вместо бега)
r = update_or_get(ws2, "2026-05-03", "утро")
ws2.cell(row=r, column=5).value = "✅ вело 60.43км / 2:19 / HR 128 (+216м)"
ws2.cell(row=r, column=6).value = "вело-длит"
ws2.cell(row=r, column=7).value = 2.32

# === Сводка ===
ws3 = wb["Сводка"]
# W1 хвост: 5.6 -> 6.7
for r in range(5, 11):
    label = ws3.cell(row=r, column=1).value
    if label and "W1" in str(label):
        ws3.cell(row=r, column=2).value = 6.7
    elif label and "ИТОГО" in str(label):
        ws3.cell(row=r, column=2).value = 63.1

# По типам — пересчёт с учётом факта W1
# Было запланировано (и в исходной сводке): вело-длит 13.3, бег 9.6 (после v2), бег-длит 7.1
# Факт W1 заменил:
#   - вело 1ч (план Пт) -> -1.0 из вело категории, и бег-длит 19км +1.85
#   - вело-длит 3.33 (план Сб) -> вело-длит остаётся, но фактическое 2.35 -> +(-0.98) к вело-длит
#     Точнее: убираем плановые 3.33, добавляем фактические 2.35 + 0.07 (перемещение Пт) + 0.15 (перемещение Сб) + 2.32 (Вс)
#     = -3.33 + 2.35 (Сб длит) + 0.07 (Пт перем) + 0.15 (Сб перем) + 2.32 (Вс) = +1.56 к вело
#     Хм, сложнее.
#
# Проще: пересчитать суммы по типам напрямую через лист «Тренировки списком».

# Пересчёт типов с нуля
type_totals = {}
for r in range(2, ws2.max_row + 1):
    typ = ws2.cell(row=r, column=6).value
    h = ws2.cell(row=r, column=7).value
    if typ and h is not None:
        type_totals[typ] = type_totals.get(typ, 0) + float(h)

print("\nПересчёт типов:")
for t, h in sorted(type_totals.items(), key=lambda x: -x[1]):
    print(f"  {t}: {h:.2f}ч ({h/63.1*100:.1f}%)")

# Очистим строки 15-30 в Сводке и переписываем заново
for r in range(15, 35):
    for c in range(1, 4):
        ws3.cell(row=r, column=c).value = None

ws3.cell(row=14, column=1).value = "Тип"
ws3.cell(row=14, column=2).value = "Часы"
ws3.cell(row=14, column=3).value = "% от месяца"

r = 15
for t, h in sorted(type_totals.items(), key=lambda x: -x[1]):
    ws3.cell(row=r, column=1).value = t
    ws3.cell(row=r, column=2).value = round(h, 2)
    ws3.cell(row=r, column=3).value = f"{h / 63.1 * 100:.1f}%"
    r += 1

wb.save(dst)
print(f"\nSaved: {dst}")
print(f"Size: {os.path.getsize(dst):,} bytes")
